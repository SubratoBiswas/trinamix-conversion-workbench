"""The Customer column order: V2_2 and V2.3 say the same thing, and V2.3 has a typo.

SESSION_HANDOFF §13.1b read as open for two days on the strength of version
numbers. The live spec named `V2 2.xlsx` as its source while the V2.3 extraction
sat unused in `docs/incoming/`, so item 1 of that entry asked for the live file to
be REPLACED by the extraction. Nobody had compared them.

Diffed 05-Aug. They are the same document in every respect that decides a load:
the same 15 interfaces in the same order, the same 15 CSV file names, 996
`fbdi_order` columns, 996 `csv_order` columns, and the same three interfaces
where the CSV order differs from the worksheet order.

There is exactly one difference, and it runs the wrong way. The extraction's last
`csv_order` entry for RA_CUSTOMER_PROFILES_INT_ALL reads

    "Review Before Consolidated Billing,"

with a trailing comma — the cell immediately before the END terminator. Oracle's
own bundled workbook says otherwise: `CustomerImport_HZ_IMP__RA_CUSTOMER.xlsm`,
sheet RA_CUSTOMER_PROFILES_INT_ALL, header row 4, column 132, no comma. Doing
what §13.1b asked would have put a column name matching nothing into a headerless
CSV, which is the failure mode this whole spec exists to prevent.

So the entry is closed WITHOUT the replacement, and these tests hold that
decision in place — including the part that would otherwise be lost, which is why
the obvious-looking action was the wrong one.
"""
import json
import os
import sys
from pathlib import Path

sys.path.insert(0, os.path.join(os.path.dirname(__file__), ".."))

_BACKEND = Path(__file__).resolve().parent.parent
_LIVE = _BACKEND / "app" / "data" / "customer_fbdi_column_order.json"
_INCOMING = (_BACKEND.parent / "docs" / "incoming"
             / "customer_fbdi_column_order_V2.json")
_TEMPLATE = (_BACKEND / "app" / "data" / "fbdi_templates"
             / "CustomerImport_HZ_IMP__RA_CUSTOMER.xlsm")

_ARTEFACT = "Review Before Consolidated Billing,"
_CORRECT = "Review Before Consolidated Billing"
_PROFILES = "RA_CUSTOMER_PROFILES_INT_ALL"
_REORDERING = {"HZ_IMP_ACCTSITES_T", "HZ_IMP_ACCTSITEUSES_T", _PROFILES}


def check(name, cond, detail=""):
    if cond:
        print(f"  PASS  {name}"); return
    raise AssertionError(f"{name} {detail}".strip())


def _live() -> dict:
    return json.loads(_LIVE.read_text(encoding="utf-8"))


# ---------------------------------------------------------------------------
# The live spec, on its own terms
# ---------------------------------------------------------------------------
def test_the_live_spec_is_the_shape_both_documents_describe():
    sheets = _live()["sheets"]
    check("15 interfaces", len(sheets) == 15, f"got {len(sheets)}")
    check("996 fbdi columns",
          sum(len(v["fbdi_order"]) for v in sheets.values()) == 996)
    check("996 csv columns",
          sum(len(v["csv_order"]) for v in sheets.values()) == 996)
    check("every interface has a CSV file name",
          all(v.get("csv") for v in sheets.values()))


def test_exactly_three_interfaces_reorder_for_the_csv():
    """CODEBASE_GUIDE §5.7, arrived at independently from the workbook. If a
    fourth ever appears, the extraction changed and this needs re-reading."""
    sheets = _live()["sheets"]
    measured = {n for n, v in sheets.items() if v["fbdi_order"] != v["csv_order"]}
    check("three reorder", measured == _REORDERING, f"got {sorted(measured)}")
    declared = {n for n, v in sheets.items() if v.get("reorders")}
    check("and the flag agrees with the measurement", declared == measured,
          f"flagged {sorted(declared)} measured {sorted(measured)}")


def test_the_live_spec_does_not_carry_the_extraction_artefact():
    """The single reason §13.1b's item 1 must not be done."""
    profiles = _live()["sheets"][_PROFILES]
    check("last csv column is clean", profiles["csv_order"][-1] == _CORRECT,
          f"got {profiles['csv_order'][-1]!r}")
    # The DATA, not the whole document. `_reconciled` quotes the artefact on
    # purpose — that is the record of what was rejected and why — so scanning the
    # file wholesale fails on its own explanation. Same trap as the three
    # source-reading tests that a comment satisfied: only the part that decides
    # a load may be searched.
    columns = json.dumps(_live()["sheets"])
    check("and no column anywhere carries it", _ARTEFACT not in columns)
    check("nor does any other column end in a comma",
          not [c for v in _live()["sheets"].values()
               for c in v["fbdi_order"] + v["csv_order"] if c.strip().endswith(",")])


def test_oracles_own_workbook_is_what_settles_it():
    """Not a preference between two internal documents — a reading of the
    template Oracle ships. Row 4 is the header row; trailing blanks are dropped
    because the sheet's declared width overstates it."""
    import openpyxl

    wb = openpyxl.load_workbook(_TEMPLATE, read_only=True, data_only=True)
    check(f"{_PROFILES} is a tab", _PROFILES in wb.sheetnames)
    row = next(wb[_PROFILES].iter_rows(min_row=4, max_row=4, values_only=True))
    header = [("" if v is None else str(v).strip()) for v in row]
    while header and not header[-1]:
        header.pop()
    check("132 columns", len(header) == 132, f"got {len(header)}")
    check("and the last one has no comma", header[-1] == _CORRECT,
          f"Oracle says {header[-1]!r}")


def test_the_spec_records_how_this_was_settled():
    """§13.1b sat open because the file said V2_2 and nothing said why that was
    still right. A version number is not provenance."""
    doc = _live()
    recon = " ".join(doc.get("_reconciled", []))
    check("the reconciliation is recorded", recon)
    check("it names the document compared against", "V2 3.xlsx" in doc["_source"]
          or "V2.3" in recon or "customer_fbdi_column_order_V2.json" in recon)
    check("it names the artefact", _ARTEFACT in recon)
    check("and says the replacement is deliberately not done",
          "not done" in recon.lower())
    check("the file carries the date of the document it agrees with",
          doc["_effective_date"] == "2026-08-03", f"got {doc['_effective_date']!r}")


# ---------------------------------------------------------------------------
# The diff itself, while both files are still on disk
# ---------------------------------------------------------------------------
def test_the_two_documents_differ_only_by_that_one_comma():
    """Re-runs the comparison rather than trusting the note about it. Skipped
    only if the incoming extraction has been archived — the checks above stand on
    the live file alone, so nothing is lost when it goes."""
    if not _INCOMING.exists():
        print(f"  SKIP  {_INCOMING.name} is no longer on disk")
        return
    live = _live()["sheets"]
    inc = json.loads(_INCOMING.read_text(encoding="utf-8"))["order"]

    check("same interfaces in the same order", list(live) == list(inc),
          f"live={list(live)} incoming={list(inc)}")
    differences = []
    for name in live:
        a, b = live[name], inc[name]
        if a["fbdi_order"] != b["fbdi_order"]:
            differences.append((name, "fbdi_order"))
        if a["csv_order"] != b["csv_order"]:
            for i, (x, y) in enumerate(zip(a["csv_order"], b["csv_order"])):
                if x != y:
                    differences.append((name, "csv_order", i, x, y))
        if (a.get("csv") or "") != (b.get("csv_file_name") or ""):
            differences.append((name, "csv name", a.get("csv"), b.get("csv_file_name")))

    check("exactly one difference", len(differences) == 1, f"got {differences}")
    name, field, index, ours, theirs = differences[0]
    check("it is the profiles interface", name == _PROFILES)
    check("in the csv order", field == "csv_order")
    check("at the last column", index == 131, f"got {index}")
    check("ours is clean", ours == _CORRECT, f"got {ours!r}")
    check("theirs carries the comma", theirs == _ARTEFACT, f"got {theirs!r}")


if __name__ == "__main__":
    for fn in [v for k, v in sorted(globals().items()) if k.startswith("test_")]:
        print(fn.__name__); fn()
    print("\nall customer spec reconciliation checks passed")
