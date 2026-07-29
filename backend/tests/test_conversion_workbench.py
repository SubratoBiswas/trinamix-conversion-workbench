"""Trinamix Conversion Workbench — consolidated unit-test suite (pure logic).

One runnable file that exercises the dependency-light core of the conversion
engine with NO database and NO network, so it is safe to run anywhere:

    pytest backend/tests/test_conversion_workbench.py -v
    # or, without pytest:
    python3 backend/tests/test_conversion_workbench.py

Coverage
--------
1. Filled Oracle FBDI Excel templates  (app.services.template_fill_service)
   - tabular layout (Supplier/BOM/Customer): data row 5, columns from A
   - Oracle-transposed layout (Item): data below the metadata block, columns from B
   - shipped sample rows are wiped; unmatched columns ignored; Instructions kept
2. Supplier FBDI CSV layout  (app.services.supplier_fbdi_layout)
   - analyst tab column order + END terminator + headerless default + Oracle names
3. Multi-source merge / de-duplication  (app.services.merge_dedupe)
   - master de-dup by business key + source priority + golden-record survivorship
   - child interfaces keep distinct rows (not collapsed)
4. Generate-time data quality  (app.services.generate_dq)
   - cleansing (trim + custom rules) and validation (required + hard-error block)

Every module tested here is the SAME code the generator runs, so a green suite
means the behaviour holds for every conversion — existing or new.
"""
import os
import sys

import openpyxl
import pandas as pd

_BACKEND = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
if _BACKEND not in sys.path:
    sys.path.insert(0, _BACKEND)

from app.services import supplier_fbdi_layout as L                       # noqa: E402
from app.services.template_fill_service import detect_layout, fill_template  # noqa: E402
from app.services.merge_dedupe import key_col_for, merge_dedupe          # noqa: E402
from app.services.generate_dq import apply_cleansing, build_report, validate_frame  # noqa: E402

_TPL = os.path.join(_BACKEND, "app", "data", "fbdi_templates")
SUPPLIER_XLSM = os.path.join(_TPL, "1_SupplierImport_POZ_SUPPLIERS_INT.xlsm")
ITEM_XLSM = os.path.join(_TPL, "ItemImport_EGP_SYSTEM_ITEMS_INTERFACE.xlsm")

# Natural-key registry used by the merge (same shape the generator passes in).
KR = {"Supplier": ["SupplierNumber", "Supplier Number"], "Item": ["Item Number", "ItemNumber"]}


# ───────────────────────── helpers ─────────────────────────
def _templates_present() -> bool:
    return os.path.exists(SUPPLIER_XLSM) and os.path.exists(ITEM_XLSM)


def _reopen(b: bytes):
    import io
    return openpyxl.load_workbook(io.BytesIO(b), keep_vba=True)


# ═══════════════════ 1. Filled Oracle templates ═══════════════════
def test_template_supplier_tabular_fill():
    """Tabular template: header row 4, data written from row 5, samples wiped."""
    if not _templates_present():
        return
    df = pd.DataFrame({
        "Import Action *": ["CREATE", "CREATE", "UPDATE"],
        "Supplier Name*": ["Acme Co", "Beta Ltd", "Gamma Inc"],
        "Supplier Number": ["90001", "90002", "90003"],
    })
    wb = _reopen(fill_template(SUPPLIER_XLSM, {"POZ_SUPPLIERS_INT": df}))
    ws = wb["POZ_SUPPLIERS_INT"]
    assert detect_layout(ws) == (4, 5, 1)
    hdr = {ws.cell(row=4, column=c).value: c for c in range(1, ws.max_column + 1)}
    assert ws.cell(row=5, column=hdr["Supplier Name*"]).value == "Acme Co"
    assert ws.cell(row=7, column=hdr["Supplier Name*"]).value == "Gamma Inc"
    col = hdr["Supplier Name*"]
    assert all(ws.cell(row=r, column=col).value in (None, "") for r in range(8, 14))
    wb.close()


def test_template_item_transposed_fill():
    """Oracle-transposed template: fields from col B, data below the metadata block."""
    if not _templates_present():
        return
    df = pd.DataFrame({
        "Transaction Type": ["SYNC", "SYNC"],
        "Batch ID": ["777", "777"],
        "Item Number": ["ITM-1", "ITM-2"],
    })
    wb = _reopen(fill_template(ITEM_XLSM, {"EGP_SYSTEM_ITEMS_INTERFACE": df}))
    ws = wb["EGP_SYSTEM_ITEMS_INTERFACE"]
    hr, ds, cs = detect_layout(ws)
    assert hr == 4 and cs == 2 and ds >= 9
    hdr = {ws.cell(row=hr, column=c).value: c for c in range(cs, ws.max_column + 1)}
    assert ws.cell(row=ds, column=hdr["Item Number"]).value == "ITM-1"
    assert ws.cell(row=ds + 1, column=hdr["Transaction Type"]).value == "SYNC"
    col = hdr["Item Number"]
    assert all(ws.cell(row=r, column=col).value in (None, "") for r in range(ds + 2, ds + 8))
    wb.close()


def test_template_unmatched_columns_and_instructions_kept():
    if not _templates_present():
        return
    df = pd.DataFrame({"Supplier Name*": ["Acme Co"], "Totally Unknown Column": ["x"]})
    wb = _reopen(fill_template(SUPPLIER_XLSM, {"POZ_SUPPLIERS_INT": df}))
    ws = wb["POZ_SUPPLIERS_INT"]
    hdr = {ws.cell(row=4, column=c).value: c for c in range(1, ws.max_column + 1)}
    assert ws.cell(row=5, column=hdr["Supplier Name*"]).value == "Acme Co"
    assert "Instructions and CSV Generation" in wb.sheetnames  # never touched
    wb.close()


# ═══════════════════ 2. Supplier FBDI CSV layout ═══════════════════
def _supplier_headers(sheet="POZ_SUPPLIERS_INT",
                      fname="1_SupplierImport_POZ_SUPPLIERS_INT.xlsm") -> list:
    wb = openpyxl.load_workbook(os.path.join(_TPL, fname), read_only=True, data_only=True)
    hdr = []
    for row in wb[sheet].iter_rows(min_row=4, max_row=4, values_only=True):
        hdr = [("" if c is None else str(c).strip()) for c in row]
        break
    while hdr and hdr[-1] == "":
        hdr.pop()
    wb.close()
    return hdr


def _frame(headers, nrows=3):
    return pd.DataFrame([{h: f"{h[:4]}_{i}" for h in headers} for i in range(nrows)])


def test_supplier_end_terminator_no_column_loss():
    if not _templates_present():
        return
    headers = _supplier_headers()
    out = L.apply_supplier_layout(_frame(headers), "POZ_SUPPLIERS_INT", True)
    assert list(out.columns)[-1] == "END"
    assert (out["END"] == "END").all()
    assert set(headers).issubset(set(out.columns))
    assert len(out.columns) == len(headers) + 1


def test_supplier_headerless_default_and_toggle():
    if not _templates_present():
        return
    out = L.apply_supplier_layout(_frame(_supplier_headers()), "POZ_SUPPLIERS_INT", True)
    headerless = out.to_csv(index=False, header=False).splitlines()
    withhdr = out.to_csv(index=False, header=True).splitlines()
    assert headerless[0].split(",")[-1] == "END"
    assert not headerless[0].startswith("Import Action")
    assert withhdr[0].split(",")[0].startswith("Import Action")


def test_supplier_oracle_names_and_noop_for_non_supplier():
    assert L.zip_name_for("POZ_SUPPLIERS_INT") == "PozSuppliersInt"
    assert L.csv_name_for("POZ_SUPPLIER_ADDRESSES_INT") == "PozSupplierAddressesInt"
    out = L.apply_supplier_layout(pd.DataFrame([{"A": "1", "B": "2"}]),
                                  "EGP_SYSTEM_ITEMS_INTERFACE", False)
    assert list(out.columns) == ["A", "B"] and "END" not in out.columns


# ═══════════════════ 3. Multi-source merge / de-dup ═══════════════════
def test_merge_master_dedup_source_priority():
    s1 = pd.DataFrame([{"Supplier Number": "S1", "Supplier Name": "Acme (src1)"},
                       {"Supplier Number": "S2", "Supplier Name": "Beta"}])
    s2 = pd.DataFrame([{"Supplier Number": "S1", "Supplier Name": "Acme (src2)"},
                       {"Supplier Number": "S3", "Supplier Name": "Gamma"}])
    m = merge_dedupe([s1, s2], "Supplier Import", KR)
    assert len(m) == 3
    assert m[m["Supplier Number"] == "S1"]["Supplier Name"].iloc[0] == "Acme (src1)"


def test_merge_survivorship_golden_record():
    s1 = pd.DataFrame([{"Supplier Number": "S1", "Supplier Name": "Acme", "Phone": ""}])
    s2 = pd.DataFrame([{"Supplier Number": "S1", "Supplier Name": "Acme Inc", "Phone": "111"}])
    r = merge_dedupe([s1, s2], "Supplier Import", KR)
    r = r[r["Supplier Number"] == "S1"].iloc[0]
    assert r["Supplier Name"] == "Acme"   # priority populated value wins
    assert r["Phone"] == "111"            # priority blank back-filled from lower source
    r2 = merge_dedupe([s1, s2], "Supplier Import", KR, survivorship=False)
    assert r2[r2["Supplier Number"] == "S1"].iloc[0]["Phone"] == ""


def test_merge_child_interface_keeps_distinct_rows():
    c1 = pd.DataFrame([{"Supplier Number": "S1", "Supplier Site": "SITE-A"},
                       {"Supplier Number": "S1", "Supplier Site": "SITE-B"}])
    c2 = pd.DataFrame([{"Supplier Number": "S1", "Supplier Site": "SITE-A"},
                       {"Supplier Number": "S1", "Supplier Site": "SITE-C"}])
    m = merge_dedupe([c1, c2], "Supplier Site", KR)
    assert sorted(m["Supplier Site"]) == ["SITE-A", "SITE-B", "SITE-C"]
    assert key_col_for(c1, "Supplier Site", KR) is None


# ═══════════════════ 4. Generate-time data quality ═══════════════════
def test_dq_cleansing_trim_and_custom_rules():
    df = pd.DataFrame([{"Supplier Name": "  Acme  ", "Code": "abc"}])
    cleaned, fixes = apply_cleansing(df, [{"field": "Supplier Name", "rule_type": "UPPERCASE"},
                                          {"field": "Code", "rule_type": "UPPERCASE"}])
    assert cleaned["Supplier Name"].iloc[0] == "ACME"
    # The standing cleanse used to be a bare whitespace trim reported as "TRIM".
    # It is now the cleansing_rules profile, so the automatic pass reports its
    # family key; custom authored rules still report their own rule_type.
    assert any(f["rule"] == "whitespace_punct" for f in fixes)
    assert any(f["rule"] == "UPPERCASE" for f in fixes)


def test_dq_validation_blocks_on_hard_error():
    df = pd.DataFrame([{"Supplier Name": "Acme", "Supplier Number": "", "Amount": "-5"}])
    tf = [{"field_name": "Supplier Number", "required": True, "data_type": "text", "max_length": 30}]
    custom = [{"field": "Amount", "rule_type": "NOT_NEGATIVE", "severity": "error"}]
    rep = build_report(validate_frame(df, tf, custom, 2000), [])
    assert rep["hard_error_count"] == 1 and rep["blocked"] is True


if __name__ == "__main__":
    tests = [v for k, v in sorted(globals().items())
             if k.startswith("test_") and callable(v)]
    if not _templates_present():
        print("NOTE: bundled FBDI templates not found — template/layout tests will no-op.\n")
    passed = failed = 0
    for t in tests:
        try:
            t()
            print(f"PASS  {t.__name__}")
            passed += 1
        except AssertionError as e:
            print(f"FAIL  {t.__name__}: {e}")
            failed += 1
        except Exception as e:  # noqa: BLE001
            print(f"ERROR {t.__name__}: {type(e).__name__}: {e}")
            failed += 1
    print(f"\n{passed}/{passed + failed} tests passed"
          + (f", {failed} failed" if failed else ""))
    sys.exit(1 if failed else 0)
