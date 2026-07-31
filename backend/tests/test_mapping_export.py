"""Unit tests for the banded field-mapping export (Issue #3)."""
import io
import os
import sys

import openpyxl

_BACKEND = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
if _BACKEND not in sys.path:
    sys.path.insert(0, _BACKEND)

from app.services.mapping_export_service import band_for, build_workbook  # noqa: E402

RECS = [
    {"target_field": "Item Number", "suggested_source": "itemid", "confidence": 100,
     "reason": "Auto-applied from learning library", "excluded": False},
    {"target_field": "Formatted Description", "suggested_source": "description", "confidence": 54,
     "reason": "column name overlap (50%)", "excluded": False,
     "alternatives": [
         {"source": "description", "confidence": 54, "verdict": "accept", "reason": "long text"},
         {"source": "notes", "confidence": 0.31, "verdict": "reject", "reason": "too short"}],
     "crosswalks": [
         {"legacy": "ACT", "oracle": "Active", "status": "vetted"},
         {"legacy": "INA", "oracle": "Inactive", "status": "unverified"}]},
    {"target_field": "Transaction Type", "suggested_source": "custitem_product_type", "confidence": 30,
     "reason": "semantic keyword match", "excluded": False},
    {"target_field": "Item Class Name", "suggested_source": "class", "confidence": 24,
     "reason": "Analyst rule: do-not-map list", "excluded": True},
    {"target_field": "Asset Tracked", "suggested_source": "createddate", "confidence": 0,
     "reason": "implausible", "excluded": False},
]


def test_the_bands_are_the_ones_the_analyst_asked_for():
    """CW_Issues row 3, verbatim: "grouped into confidence bands (eg. 90-100, 80-90,
    75-80, 50-75, below 50, 0)". This shipped with six bands of its own invention at
    different cut points — the same idea, close enough to look done, and impossible
    for the functional team to reconcile against what they asked for."""
    from app.services.mapping_export_service import BANDS
    labels = [lbl for _k, lbl, *_ in BANDS]
    assert labels == ["100% - Exact Match", "90-100%", "80-90%", "75-80%",
                      "50-75%", "Below 50%", "0% - No Match Found"], labels


def test_band_for_boundaries():
    assert band_for(100) == "exact"
    # Every boundary, on both sides — a band table is only as good as its edges.
    assert band_for(99) == "b90" and band_for(90) == "b90"
    assert band_for(89) == "b80" and band_for(80) == "b80"
    assert band_for(79) == "b75" and band_for(75) == "b75"
    assert band_for(74) == "b50" and band_for(50) == "b50"
    assert band_for(49) == "b0" and band_for(1) == "b0"
    assert band_for(0) == "none" and band_for(None) == "none"


def test_no_confidence_falls_in_a_band():
    """Every 0-100 value lands in exactly one band — no gaps, no overlaps. A gap
    would silently drop fields out of the workbook entirely."""
    from app.services.mapping_export_service import BANDS
    keys = {k for k, *_ in BANDS}
    for c in range(0, 101):
        assert band_for(c) in keys, c


def _open():
    return openpyxl.load_workbook(io.BytesIO(build_workbook("Item to NetSuite Field Mapping", RECS)))


def test_summary_counts_and_total():
    wb = _open()
    s = wb["Summary"]
    counts = {r[0]: r[1] for r in s.iter_rows(min_row=4, max_row=12, values_only=True) if r[0]}
    assert counts["100% - Exact Match"] == 1
    assert counts["Below 50%"] == 2
    assert counts["0% - No Match Found"] == 1
    assert counts["Total"] == 5


def test_only_nonempty_bands_get_sheets():
    wb = _open()
    assert "100pct_-_Exact_Match" in wb.sheetnames
    assert "Below_50pct" in wb.sheetnames
    assert "90-100pct" not in wb.sheetnames  # empty band → no sheet


def test_band_sheet_headers_and_rows():
    wb = _open()
    ws = wb["Below_50pct"]
    assert [ws.cell(row=2, column=c).value for c in range(1, 8)] == [
        "Target FBDI Field", "Suggested Source Field", "Confidence %", "Reason",
        "Excluded (Do-Not-Map Rule)", "Vetted Alternatives (AI-checked)",
        "Value Crosswalks (legacy → Oracle)"]
    # excluded row renders "Yes"
    vals = [[ws.cell(row=r, column=c).value for c in range(1, 6)] for r in range(3, 5)]
    excl = {tuple(v[:2]): v[4] for v in vals}
    assert excl[("Item Class Name", "class")] == "Yes"
    assert excl[("Transaction Type", "custitem_product_type")] in (None, "")


def test_vetted_alternatives_and_crosswalks_render():
    wb = _open()
    ws = wb["50-75pct"]  # the Formatted Description row lives here
    # find the row for Formatted Description
    row = next(r for r in range(3, ws.max_row + 1)
               if ws.cell(row=r, column=1).value == "Formatted Description")
    alts = ws.cell(row=row, column=6).value or ""
    cws = ws.cell(row=row, column=7).value or ""
    # ranked alternatives, normalised to % with AI verdict + reason
    assert "description (54%) — accept: long text" in alts
    assert "notes (31%) — reject: too short" in alts       # 0.31 → 31%
    # value crosswalks legacy → Oracle with status
    assert "ACT → Active  (vetted)" in cws
    assert "INA → Inactive  (unverified)" in cws


def test_missing_vetted_fields_are_blank():
    wb = _open()
    ws = wb["100pct_-_Exact_Match"]  # Item Number row has no alts/crosswalks
    assert (ws.cell(row=3, column=6).value or "") == ""
    assert (ws.cell(row=3, column=7).value or "") == ""


def test_title_on_summary():
    wb = _open()
    assert wb["Summary"]["A1"].value == "Item to NetSuite Field Mapping"


if __name__ == "__main__":
    tests = [v for k, v in sorted(globals().items()) if k.startswith("test_") and callable(v)]
    p = 0
    for t in tests:
        t(); print("PASS ", t.__name__); p += 1
    print(f"\n{p}/{len(tests)} passed")
