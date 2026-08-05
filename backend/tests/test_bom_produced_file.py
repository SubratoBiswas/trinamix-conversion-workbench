"""The BOM CSVs Oracle actually receives — checked against the spec, column by column.

WHY THIS EXISTS
---------------
``test_bom_fbdi_sequence.py`` tests the layout FUNCTION and asserts that
``output_service`` calls it. Both pass. Neither had ever looked at a produced
file, and BOM_AND_USERS_HANDOFF §1 asks for exactly that before anyone loads one:

    "Verify the generated BOM CSVs column-by-column against this JSON before
     anyone loads them."

The gap was not academic. Nothing in the suite CALLED ``generate_output_artifact``,
and on 04-Aug it acquired a log line that read ``obj_name`` sixty lines before the
name was assigned — ``UnboundLocalError`` on every generate, every format, every
object, with 1,182 tests green over it. A test that produces a file is the only
kind that could have said so.

WHAT IS CHECKED, AND WHY EACH ONE MATTERS
-----------------------------------------
These are HEADERLESS CSVs. Column POSITION is the only thing carrying meaning, so
a file with the right names in the wrong order loads silently into the wrong
fields and nothing downstream objects.

  1. Four files, named exactly as Oracle names them. Oracle matches by file name,
     so a numbered prefix — which the supplier and customer packages both add —
     would be rejected.
  2. Header row identical to the spec, position for position, on all four.
  3. No END column. The supplier package appends a record terminator and BOM
     reuses that machinery; inheriting it adds a field Oracle does not expect.
  4. The headerless file is the same width as the spec.
  5. A distinctive source value lands at the INDEX the spec gives for the field
     it was mapped to. This is the check that cannot be satisfied by accident:
     it reads the file the way the loader reads it, by position.

AND THE SPEC ITSELF IS CHECKED AGAINST ORACLE'S OWN TEMPLATE.
``bom_fbdi_column_order.json`` was transcribed from the analyst's mapping
workbook. The bundled ``BOMItemStructure_EGP_STRUCTURES_INT.xlsm`` is Oracle's,
carries all four interfaces, and is an independent witness — so the chain runs
Oracle's template → the spec → the produced file, with no step taken on trust.

THE IN-MEMORY DATABASE
----------------------
``mongomock_motor`` backs Beanie with a real-enough Mongo in-process, which is
what makes an end-to-end generate runnable in a suite that must not require
infrastructure. It is in requirements.txt for that reason. If it is ever absent
the module skips rather than fails — but a skip here means the only test that
opens a produced file is not running, so treat a skip as a red flag rather than
as noise.
"""
import csv
import io
import json
import os
import sys
import tempfile
import zipfile
from pathlib import Path

import pytest

sys.path.insert(0, os.path.join(os.path.dirname(__file__), ".."))

pytest.importorskip(
    "mongomock_motor",
    reason="mongomock_motor is not installed — the produced-file BOM check cannot run",
)

import pandas as pd                                              # noqa: E402
from beanie import init_beanie                                   # noqa: E402
from mongomock_motor import AsyncMongoMockClient                 # noqa: E402

from app.config import settings                                  # noqa: E402
from app.models.client import Client                             # noqa: E402
from app.models.conversion import Conversion                     # noqa: E402
from app.models.dataset import Dataset, DatasetColumnProfile     # noqa: E402
from app.models.fbdi import (                                    # noqa: E402
    FBDIField, FBDISheet, FBDITemplate, FBDITemplateFile, GoldStandard, OracleLookup,
)
from app.models.mapping import MappingSuggestion                 # noqa: E402
from app.models.output import ConvertedOutput                    # noqa: E402
from app.models.project import Project                           # noqa: E402
from app.models.transformation import Crosswalk, TransformationRule  # noqa: E402
from app.models.user import User                                 # noqa: E402

_BACKEND = Path(__file__).resolve().parent.parent
_TEMPLATE = (_BACKEND / "app" / "data" / "fbdi_templates"
             / "BOMItemStructure_EGP_STRUCTURES_INT.xlsm")
_SPEC = json.loads((_BACKEND / "app" / "data"
                    / "bom_fbdi_column_order.json").read_text(encoding="utf-8"))
_ORDER = _SPEC["order"]
_CSV_NAMES = _SPEC["csv_file_names"]
_IFACE_BY_FILE = {v: k for k, v in _CSV_NAMES.items()}

# One distinctive value per mapped source column, so a misplaced column shows up
# as a value at the wrong index rather than as a subtle difference in a name.
_PROBES = {"STRUCT": "PROBE-STRUCTURE", "ORG": "PROBE-ORG", "ITEM": "PROBE-ITEM",
           # The detail columns the Substitutes and Reference Designators tabs are
           # now FILTERED on — a line without them is dropped from those tabs (BOM
           # feedback, 05-Aug), so the probe rows must carry them or those tabs
           # come back empty.
           "SUBST": "PROBE-SUBSTITUTE", "REFDES": "PROBE-REFDES"}

# Which template field each probe column is mapped into, per interface. Only
# fields every one of these interfaces genuinely has.
_MAPPED = {
    "EGP_STRUCTURES_INTERFACE": {"Structure Name": "STRUCT",
                                 "Organization Code": "ORG",
                                 "Item Name": "ITEM"},
    "EGP_COMPONENTS_INTERFACE": {"Structure Name": "STRUCT", "Organization Code": "ORG"},
    "EGP_SUB_COMPS_INTERFACE": {"Structure Name": "STRUCT", "Organization Code": "ORG",
                                "Substitute Item Name": "SUBST"},
    "EGP_REF_DESGS_INTERFACE": {"Structure Name": "STRUCT", "Organization Code": "ORG",
                                "Reference Designator": "REFDES"},
}


def check(name, cond, detail=""):
    if cond:
        print(f"  PASS  {name}"); return
    raise AssertionError(f"{name} {detail}".strip())


# ---------------------------------------------------------------------------
# The harness: a whole conversion, in memory, generated once and reused.
# ---------------------------------------------------------------------------
_MODELS = [
    User, Client, Project, Conversion, Dataset, DatasetColumnProfile,
    FBDITemplate, FBDISheet, FBDIField, FBDITemplateFile, OracleLookup,
    GoldStandard, MappingSuggestion, TransformationRule, Crosswalk, ConvertedOutput,
]


async def _build_conversion(tmp: Path):
    from app.parsers import parse_fbdi_template

    parsed = parse_fbdi_template(_TEMPLATE)
    template = FBDITemplate(name="BOM Item Structure", business_object="BOM",
                            file_name=_TEMPLATE.name, status="parsed", is_global=True)
    await template.insert()

    sheets = {}
    for s in parsed["sheets"]:
        sheet = FBDISheet(template_id=template.id, sheet_name=s["sheet_name"],
                          sequence=s.get("sequence", 0),
                          field_count=s.get("field_count", 0))
        await sheet.insert()
        sheets[s["sheet_name"]] = sheet

    fields: dict[str, list] = {}
    for f in parsed["fields"]:
        sheet = sheets[f["sheet_name"]]
        field = FBDIField(template_id=template.id, sheet_id=sheet.id,
                          field_name=f["field_name"], display_name=f.get("display_name"),
                          required=f.get("required", False), data_type=f.get("data_type"),
                          sequence=f.get("sequence", 0))
        await field.insert()
        fields.setdefault(f["sheet_name"], []).append(field)

    source = tmp / "bom_source.csv"
    pd.DataFrame({k: [v, v + "-2"] for k, v in _PROBES.items()}).to_csv(source, index=False)

    client = Client(name="Probe", code="PRB", is_default=True); await client.insert()
    project = Project(name="BOM produced-file check", client_id=client.id)
    await project.insert()
    dataset = Dataset(name="bom source", file_name=source.name, file_path=str(source),
                      file_type="csv", row_count=2, column_count=len(_PROBES))
    await dataset.insert()
    conversion = Conversion(project_id=project.id, name="BOM produced-file check",
                            dataset_id=dataset.id, dataset_ids=[dataset.id],
                            template_id=template.id, target_object="BOM",
                            source_type="dataset")
    await conversion.insert()

    for sheet_name, wanted in _MAPPED.items():
        by_name = {f.field_name: f for f in fields[sheet_name]}
        for field_name, probe in wanted.items():
            field = by_name.get(field_name)
            check(f"{sheet_name} has {field_name}", field is not None)
            await MappingSuggestion(conversion_id=conversion.id,
                                    target_field_id=field.id, source_column=probe,
                                    confidence=1.0, status="approved").insert()
    return conversion


_GENERATED: dict = {}


def _generated() -> dict:
    """Generate once: ``{"with_header": {file: rows}, "headerless": {...}}``.

    Both formats come from the SAME conversion. The header run is how the column
    names are read at all — the shipped file has none — and the headerless run is
    what Oracle is actually given, so the checks below never assume one stands in
    for the other.
    """
    if _GENERATED:
        return _GENERATED
    import asyncio

    async def run():
        tmp = Path(tempfile.mkdtemp(prefix="bom_produced_"))
        settings.UPLOAD_DIR = str(tmp)
        settings.OUTPUT_DIR = str(tmp / "outputs")
        await init_beanie(database=AsyncMongoMockClient()["bom_produced"],
                          document_models=_MODELS)
        conversion = await _build_conversion(tmp)
        from app.services.output_service import generate_output_artifact

        out = {}
        for key, header in (("with_header", True), ("headerless", False)):
            artifact = await generate_output_artifact(conversion, fmt="csv",
                                                      include_header=header)
            path = Path(artifact.output_file_path)
            check(f"{key} is a zip package", path.suffix == ".zip", f"got {path.name}")
            with zipfile.ZipFile(path) as zf:
                out[key] = {
                    name: list(csv.reader(io.StringIO(
                        zf.read(name).decode("utf-8-sig"))))
                    for name in zf.namelist()
                }
        return out

    _GENERATED.update(asyncio.run(run()))
    return _GENERATED


# ---------------------------------------------------------------------------
# The spec against Oracle's own template — the independent witness
# ---------------------------------------------------------------------------
def test_the_spec_matches_oracles_own_bom_template_column_for_column():
    """``bom_fbdi_column_order.json`` was transcribed from the analyst's mapping
    workbook. This is the first time it has been read back against Oracle's, which
    is the only copy that decides whether a load succeeds.

    Row 4 is the header row on every tab; ``max_column`` overstates the width
    (stray formatting reaches column 78 on SUB_COMPS), so the row is read and
    trailing blanks dropped rather than trusting the sheet dimensions.
    """
    import openpyxl

    wb = openpyxl.load_workbook(_TEMPLATE, read_only=True, data_only=True)
    for iface, want in _ORDER.items():
        check(f"{iface} is a tab in Oracle's template", iface in wb.sheetnames)
        row = next(wb[iface].iter_rows(min_row=4, max_row=4, values_only=True))
        got = [("" if v is None else str(v).strip()) for v in row]
        while got and not got[-1]:
            got.pop()
        check(f"{iface} width", len(got) == len(want), f"template {len(got)} vs spec {len(want)}")
        for i, (a, b) in enumerate(zip(got, [str(c).strip() for c in want]), 1):
            check(f"{iface} col {i}", a == b, f"template={a!r} spec={b!r}")


# ---------------------------------------------------------------------------
# The produced files
# ---------------------------------------------------------------------------
def test_the_package_holds_exactly_the_four_interfaces_under_oracles_names():
    """Oracle matches the file inside the zip by NAME. The supplier and customer
    packages both prefix an ordering number onto theirs; BOM must not."""
    names = sorted(_generated()["headerless"])
    check("four files", len(names) == 4, f"got {names}")
    check("named as the spec names them", names == sorted(_CSV_NAMES.values()),
          f"got {names}")
    for n in names:
        check(f"{n} carries no ordering prefix", not n[:1].isdigit(), f"got {n}")


def test_every_produced_column_is_the_spec_column_in_the_spec_position():
    """The check the handoff asks for, on the file rather than on the function."""
    for name, rows in _generated()["with_header"].items():
        iface = _IFACE_BY_FILE[name]
        want = [str(c).strip() for c in _ORDER[iface]]
        got = [c.strip() for c in rows[0]]
        check(f"{name} width", len(got) == len(want),
              f"produced {len(got)} vs spec {len(want)}")
        for i, (a, b) in enumerate(zip(got, want), 1):
            check(f"{name} col {i}", a == b, f"produced={a!r} spec={b!r}")


def test_no_produced_file_carries_an_end_terminator():
    """The supplier package appends END and BOM reuses that machinery. A stray END
    shifts nothing but hands Oracle a field it does not expect on these four."""
    for name, rows in _generated()["with_header"].items():
        upper = [c.strip().upper() for c in rows[0]]
        check(f"{name} has no END column", "END" not in upper)
    for name, rows in _generated()["headerless"].items():
        for r, row in enumerate(rows, 1):
            check(f"{name} row {r} has no END cell",
                  "END" not in [c.strip().upper() for c in row])


def test_the_shipped_file_is_headerless_and_the_right_width():
    """A header line is read by the FBDI loader as data and rejects the first
    record, so the shipped CSV must start at row one with real values."""
    for name, rows in _generated()["headerless"].items():
        iface = _IFACE_BY_FILE[name]
        want = len(_ORDER[iface])
        check(f"{name} has rows", len(rows) >= 1)
        widths = {len(r) for r in rows}
        check(f"{name} every row is {want} wide", widths == {want}, f"got {widths}")
        first = [c.strip() for c in rows[0]]
        check(f"{name} does not start with the header",
              first != [str(c).strip() for c in _ORDER[iface]])


def test_a_mapped_value_lands_at_the_index_the_spec_gives_it():
    """The one check that cannot pass by accident. Reads the file the way the
    loader reads it — by position — and asks whether the value the analyst mapped
    into "Organization Code" is sitting where the spec says that column is."""
    for name, rows in _generated()["headerless"].items():
        iface = _IFACE_BY_FILE[name]
        order = [str(c).strip() for c in _ORDER[iface]]
        row = rows[0]
        for field_name, probe in _MAPPED[iface].items():
            value = _PROBES[probe]
            want_index = order.index(field_name)
            at = [i for i, v in enumerate(row) if v.strip() == value]
            check(f"{name}: {value} appears once", len(at) == 1,
                  f"found at {at}")
            check(f"{name}: {field_name} is at position {want_index}",
                  at[0] == want_index,
                  f"{value} landed at {at[0]} = {order[at[0]]!r}")


def test_generation_actually_returns_a_file_at_all():
    """Stated separately and on purpose. This is the assertion that was missing on
    04-Aug: every other test in the suite passed while this function could not
    reach its first write."""
    for key in ("with_header", "headerless"):
        files = _generated()[key]
        check(f"{key} produced files", bool(files))
        check(f"{key} produced rows",
              all(len(rows) >= 1 for rows in files.values()))


if __name__ == "__main__":
    for fn in [v for k, v in sorted(globals().items()) if k.startswith("test_")]:
        print(fn.__name__); fn()
    print("\nall produced-file BOM checks passed")
