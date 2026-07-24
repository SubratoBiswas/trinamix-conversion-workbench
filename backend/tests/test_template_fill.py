"""Unit tests for filling the real Oracle FBDI Excel templates.

Loads the bundled templates and verifies the fill writer places data in the right
columns/rows for both layouts (tabular supplier, Oracle-transposed item) and wipes
the shipped sample rows.
"""
import io
from pathlib import Path

import openpyxl
import pandas as pd
import pytest

from app.services.template_fill_service import fill_template, detect_layout

_DIR = Path(__file__).resolve().parent.parent / "app" / "data" / "fbdi_templates"
SUPPLIER = _DIR / "1_SupplierImport_POZ_SUPPLIERS_INT.xlsm"
ITEM = _DIR / "ItemImport_EGP_SYSTEM_ITEMS_INTERFACE.xlsm"

pytestmark = pytest.mark.skipif(
    not SUPPLIER.exists() or not ITEM.exists(),
    reason="bundled FBDI templates not present",
)


def _open(b: bytes):
    return openpyxl.load_workbook(io.BytesIO(b), keep_vba=True)


def test_supplier_tabular_layout_and_fill():
    df = pd.DataFrame({
        "Import Action *": ["CREATE", "CREATE", "UPDATE"],
        "Supplier Name*": ["Acme Co", "Beta Ltd", "Gamma Inc"],
        "Supplier Number": ["90001", "90002", "90003"],
    })
    b = fill_template(SUPPLIER, {"POZ_SUPPLIERS_INT": df})
    wb = _open(b)
    ws = wb["POZ_SUPPLIERS_INT"]
    hr, ds, cs = detect_layout(ws)
    assert (hr, ds, cs) == (4, 5, 1)  # title/req/header rows, data from 5, col A
    hdr = {ws.cell(row=hr, column=c).value: c for c in range(1, ws.max_column + 1)}
    # row 5 = first data row
    assert ws.cell(row=5, column=hdr["Supplier Name*"]).value == "Acme Co"
    assert ws.cell(row=5, column=hdr["Import Action *"]).value == "CREATE"
    assert ws.cell(row=7, column=hdr["Supplier Name*"]).value == "Gamma Inc"
    # shipped sample rows below our 3 must be cleared
    col = hdr["Supplier Name*"]
    assert all(ws.cell(row=r, column=col).value in (None, "")
               for r in range(8, 14))
    wb.close()


def test_item_transposed_layout_and_fill():
    df = pd.DataFrame({
        "Transaction Type": ["SYNC", "SYNC"],
        "Batch ID": ["777", "777"],
        "Item Number": ["ITM-1", "ITM-2"],
    })
    b = fill_template(ITEM, {"EGP_SYSTEM_ITEMS_INTERFACE": df})
    wb = _open(b)
    ws = wb["EGP_SYSTEM_ITEMS_INTERFACE"]
    hr, ds, cs = detect_layout(ws)
    # header on the 'Name' row (4), data below the metadata block, fields from col B
    assert hr == 4 and cs == 2 and ds >= 9
    hdr = {ws.cell(row=hr, column=c).value: c for c in range(cs, ws.max_column + 1)}
    assert ws.cell(row=ds, column=hdr["Item Number"]).value == "ITM-1"
    assert ws.cell(row=ds + 1, column=hdr["Transaction Type"]).value == "SYNC"
    col = hdr["Item Number"]
    assert all(ws.cell(row=r, column=col).value in (None, "")
               for r in range(ds + 2, ds + 8))
    wb.close()


def test_unmatched_columns_are_ignored():
    # A df column with no matching template header must not raise or misplace data.
    df = pd.DataFrame({
        "Supplier Name*": ["Acme Co"],
        "Totally Unknown Column": ["x"],
    })
    b = fill_template(SUPPLIER, {"POZ_SUPPLIERS_INT": df})
    wb = _open(b)
    ws = wb["POZ_SUPPLIERS_INT"]
    hdr = {ws.cell(row=4, column=c).value: c for c in range(1, ws.max_column + 1)}
    assert ws.cell(row=5, column=hdr["Supplier Name*"]).value == "Acme Co"
    wb.close()


def test_instructions_sheet_untouched():
    df = pd.DataFrame({"Supplier Name*": ["Acme Co"]})
    b = fill_template(SUPPLIER, {"POZ_SUPPLIERS_INT": df})
    wb = _open(b)
    assert "Instructions and CSV Generation" in wb.sheetnames
    wb.close()
