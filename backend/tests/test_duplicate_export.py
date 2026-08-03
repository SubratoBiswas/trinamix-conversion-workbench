""""242 more are not listed and cannot be decided here … raise the limit to review
them all" — the panel's own banner, next to no control that raised it.

The scan returns at most ``max_clusters`` groups. On the NextPower supplier run
that was 100 of 342, and the counters described the visible slice only. The
banner was honest about it and then asked the analyst to do something the screen
could not do, which reads as their oversight rather than as a missing feature.

Two answers, tested here. The panel can now ask for all of them — the API always
accepted up to 2000, only the client hardcoded 100. And every group can be pulled
as a spreadsheet, because 342 groups over 831 records is not a list anyone reviews
by scrolling, and the question actually asked of it — WHICH GROUPS DISAGREE ON A
STRONG IDENTIFIER, and are therefore probably not one entity — is a filter and a
sort.

The live case that prompted it: nine "Flextronics Technologies India" rows at 100%
name match carrying three different Tax Registration Numbers, four D-U-N-S numbers
and nine Supplier Numbers. Merging keeps ONE value per column and discards the
rest, so whether that group is one entity is the whole question, and it is not
answerable while 242 of its siblings are invisible.

Pure: hand-built scan results through the builder, plus openpyxl to read the bytes
back. No database.
"""
from io import BytesIO
from pathlib import Path

import openpyxl

from app.services import duplicate_export_service as des

_BACKEND = Path(__file__).resolve().parent.parent

NAME = "Supplier Name"
TAX = "Tax Registration Number"
DUNS = "D-U-N-S Number"
NUM = "Supplier Number"
FLEX = "FLEXTRONICS TECHNOLOGIES INDIA PVT LTD"


def member(row, name=FLEX, tax="", duns="", num=""):
    return {"row": row, "values": {NAME: name, TAX: tax, DUNS: duns, NUM: num}}


# The group from the screenshot, values as they appeared.
FLEXTRONICS = {
    "confidence": 1.0, "size": 9, "fields": [NAME, TAX],
    "members": [
        member(1416, num="172654"),
        member(1417, tax="AAACF5248E EM009", num="3604053"),
        member(1418, tax="AAACF5248E EM009", num="3604159"),
        member(1419, duns="172654", num="3792632"),
        member(1692, name="Flextronics Technologies India Pvt Ltd",
               duns="NXT_ind_pvt_vendor_159", num="1664912"),
        member(3182, name="Flextronics Technologies India Private Limited",
               tax="33AAACF5248E1Z0", num="2453370"),
        member(3807, name="Flextronics Technologies (India) Private Limited",
               duns="NXT_ind_pvt_vendor_345", num="1665098"),
        member(3808, name="FLEXTRONICS TECHNOLOGIES INDIA PRIVATE LIMITED",
               tax="27AAACF5248E3ZR", duns="NXT_ind_pvt_vendor_347", num="1665100"),
        member(3812, name="FLEXTRONICS TECHNOLOGIES India PVT LTD",
               duns="NXT_ind_pvt_vendor_350", num="1665103"),
    ],
    "id_conflicts": [
        {"column": TAX, "values": ["AAACF5248E EM009", "33AAACF5248E1Z0",
                                   "27AAACF5248E3ZR"]},
        {"column": NUM, "values": ["172654", "3604053", "3604159"]},
    ],
    "decision": None,
}

# A group nobody has flagged: same company, no disagreement on a strong id.
CLEAN = {
    "confidence": 0.94, "size": 2, "fields": [NAME],
    "members": [member(11, "PMEA SOLAR TECH SOLUTIONS LIMITED", num="3482068"),
                member(12, "PMEA SOLAR TECH SOLUTIONS LIMITED", num="3482068")],
    "id_conflicts": [],
    "decision": {"verdict": "merge", "survivor_key": "k1"},
}

RESULT = {
    "object": "Supplier", "rows_scanned": 3813, "cluster_count": 342,
    "identity_fields": [NAME, TAX, DUNS, NUM],
    "clusters": [FLEXTRONICS, CLEAN],
    "coverage_note": "23 row(s) have no value in 'Supplier Name'",
    "sources": ["NXT Supplier Extract.xlsx"],
    "ai_used": False,
}


def _wb(result=RESULT):
    return openpyxl.load_workbook(BytesIO(des.build_workbook(
        title="Supplier Import", generated_at="03 Aug 2026 19:40 UTC",
        result=result)))


def _cells(ws):
    return [[c.value for c in row] for row in ws.iter_rows()]


def _flat(ws):
    return "\n".join(" | ".join("" if v is None else str(v) for v in row)
                     for row in _cells(ws))


# ── The conflict is the point of the file ────────────────────────────────────

def test_the_conflicting_ids_are_named_with_their_actual_values():
    """Not "has conflicts" — the values. "Tax Registration Number: A, B, C" is
    what lets somebody see three GSTINs sharing one PAN and decide; a boolean
    sends them back to the screen they were trying to leave."""
    txt = des.conflict_text(FLEXTRONICS)
    assert TAX in txt
    for v in ("AAACF5248E EM009", "33AAACF5248E1Z0", "27AAACF5248E3ZR"):
        assert v in txt, v


def test_a_group_that_agrees_reports_an_empty_conflict_not_the_word_none():
    """Blank filters and sorts as the absence it is. A word has to be read."""
    assert des.conflict_text(CLEAN) == ""


def test_a_column_with_only_one_distinct_value_is_not_a_conflict():
    """One value repeated is agreement, and calling it a disagreement would put
    every group on the list the analyst is trying to narrow."""
    assert des.conflict_text({"id_conflicts": [{"column": TAX, "values": ["X"]}]}) == ""
    assert des.conflict_text({"id_conflicts": [{"column": TAX, "values": ["X", ""]}]}) == ""


def test_the_conflicts_sheet_lists_one_row_per_disagreeing_column():
    ws = _wb()["Conflicting IDs"]
    txt = _flat(ws)
    assert TAX in txt and NUM in txt
    assert "27AAACF5248E3ZR" in txt
    # The clean group contributes nothing.
    assert "PMEA" not in txt


def test_the_conflicts_sheet_says_so_plainly_when_there_are_none():
    ws = _wb({**RESULT, "clusters": [CLEAN]})["Conflicting IDs"]
    assert "No group disagrees on a strong identifier." in _flat(ws)


# ── Counts describe what is IN the file ──────────────────────────────────────

def test_the_summary_separates_what_was_found_from_what_is_listed():
    """The defect being answered. Reporting only one number is how "0 undecided"
    came to sit beside "342 suspected groups"."""
    s = des.summarise(RESULT)
    assert s["groups_found"] == 342
    assert s["groups_listed"] == 2
    assert s["records_in_groups"] == 11


def test_the_summary_counts_the_conflicted_groups_separately():
    s = des.summarise(RESULT)
    assert s["groups_with_conflicting_ids"] == 1
    assert s["records_in_conflicted_groups"] == 9


def test_a_decided_group_is_counted_as_decided():
    assert des.summarise(RESULT)["groups_decided"] == 1


def test_the_summary_carries_the_scan_coverage_note():
    """"No duplicates found" and "most of these were never compared" must not
    render identically — the same rule the panel already follows."""
    assert "23 row(s)" in _flat(_wb()["Summary"])


def test_the_summary_says_whether_ai_adjudicated():
    assert "deterministic" in _flat(_wb()["Summary"]).lower()


# ── The groups sheet ─────────────────────────────────────────────────────────

def test_every_member_of_every_group_gets_a_row():
    ws = _wb()["Groups"]
    rows = [r for r in _cells(ws)[4:] if r and r[6] not in (None, "")]
    assert len(rows) == 11, [r[6] for r in rows]


def test_the_identity_values_are_carried_so_the_file_stands_alone():
    """A group list that makes you go back to the extract to see the names is a
    list of row numbers."""
    txt = _flat(_wb()["Groups"])
    for v in (FLEX, "3792632", "NXT_ind_pvt_vendor_347", "PMEA SOLAR TECH SOLUTIONS LIMITED"):
        assert v in txt, v


def test_the_group_header_row_carries_the_match_and_the_conflict():
    ws = _wb()["Groups"]
    first = _cells(ws)[4]
    assert first[1] == "100%"
    assert first[2] == 9
    assert TAX in str(first[4])


def test_the_match_is_stated_once_per_group_not_once_per_row():
    """Repeating it down nine rows reads as nine separate findings."""
    ws = _wb()["Groups"]
    body = [r for r in _cells(ws)[4:] if r and r[6] not in (None, "")]
    assert [r[1] for r in body].count("100%") == 1


def test_a_recorded_verdict_is_spelled_out_rather_than_left_as_a_keyword():
    """"keep_survivor" is a database value. The person reading this did not
    write it."""
    txt = _flat(_wb()["Groups"])
    assert "Merge into one golden record" in txt
    assert "keep_survivor" not in txt


def test_an_undecided_group_leaves_the_decision_blank():
    """Blank means nobody has ruled and every record still ships — which is the
    state 242 of these groups were in without anyone being able to see it."""
    ws = _wb()["Groups"]
    assert not str(_cells(ws)[4][5] or "").strip()


def test_an_empty_scan_still_produces_a_readable_file():
    """A file that says "nothing found" is an answer. A file with a header row
    and nothing under it looks like a failed download."""
    ws = _wb({**RESULT, "clusters": [], "cluster_count": 0})["Groups"]
    assert "No suspected duplicate groups." in _flat(ws)


def test_the_three_sheets_are_there_and_named_for_what_they_hold():
    assert _wb().sheetnames == ["Summary", "Groups", "Conflicting IDs"]


def test_the_builder_is_pure_bytes_and_touches_nothing():
    src = (_BACKEND / "app" / "services" / "duplicate_export_service.py").read_text(
        encoding="utf-8")
    for forbidden in ("await ", "import pandas", "find_duplicate_clusters",
                      "Document", "motor"):
        assert forbidden not in src, forbidden


# ── The caller calls it ──────────────────────────────────────────────────────

_OPS = (_BACKEND / "app" / "routers" / "operations.py").read_text(encoding="utf-8")


def test_the_export_endpoint_exists_and_streams_a_workbook():
    assert '"/{conversion_id}/duplicate-suspects-export"' in _OPS
    body = _OPS.split("async def duplicate_suspects_export(")[1].split("\n@")[0]
    assert "des.build_workbook(" in body
    assert "StreamingResponse(" in body


def test_the_export_asks_for_every_group_not_the_screens_hundred():
    """Exporting the same truncated slice would reproduce the problem in a second
    format."""
    body = _OPS.split("async def duplicate_suspects_export(")[1].split("\n@")[0]
    assert "max_clusters: int = Query(2000" in body


def test_the_export_and_the_panel_run_the_SAME_scan():
    """The export exists to show the groups the panel truncates. One that scanned
    differently would answer a different question from the screen it completes —
    and a second copy of a rule drifting from the first is this codebase's most
    repeated defect."""
    assert "async def _scan_duplicates(" in _OPS
    for caller in ("duplicate_candidates", "duplicate_suspects_export"):
        body = _OPS.split(f"async def {caller}(")[1].split("\n@")[0]
        assert "_scan_duplicates(" in body, caller


def test_the_export_does_not_bake_in_an_ai_verdict():
    """AI adjudication is a decision aid on screen. Silently baked into a file
    somebody will act on, it becomes an unattributed judgement."""
    body = _OPS.split("async def duplicate_suspects_export(")[1].split("\n@")[0]
    assert "use_ai=False" in body


# ── The panel can now load them all ──────────────────────────────────────────

_PAGE = (_BACKEND.parent / "frontend" / "src" / "pages" / "OutputPreviewPage.tsx")
_API = (_BACKEND.parent / "frontend" / "src" / "api" / "index.ts")


def test_the_whole_list_is_requested_not_the_screens_hundred():
    """The client asked for 100 and the panel then told the analyst to raise a
    limit it gave them no way to raise. Asking for everything is cheap — the
    scanner computes every cluster and then SLICES, so the cap saved no work."""
    if not _API.exists():
        return
    src = _API.read_text(encoding="utf-8")
    assert "max_clusters: opts?.maxClusters ?? 2000," in src
    assert "?? 100," not in src.split("duplicateCandidates:")[1][:1200]


def test_the_requested_limit_stays_within_what_the_endpoint_accepts():
    """Asking for more than the endpoint allows is a 422, which the analyst reads
    as a broken page."""
    assert "max_clusters: int = Query(100, ge=10, le=2000)" in _OPS
    if not _PAGE.exists():
        return
    assert "const DUPE_LIMIT = 2000;" in _PAGE.read_text(encoding="utf-8")


def test_the_list_is_paged_but_the_counts_are_not():
    """THE BUG THE PAGING MUST NOT REINTRODUCE. Rendering 342 group tables at once
    is what the old cap was really protecting against — and it protected against
    it by hiding the data and then reporting on what was left. So the page slices
    the RENDER and nothing else: every count, and every bulk action, reads the
    whole list."""
    if not _PAGE.exists():
        return
    src = _PAGE.read_text(encoding="utf-8")
    # The render is the paged slice...
    assert "dupPageClusters.map(renderCluster)" in src
    assert "dupes.clusters.map(renderCluster)" not in src
    # ...and the merge target is not.
    target = src.split("const highConfidenceTargets = useMemo(")[1].split(");")[0]
    assert "dupes?.clusters ?? []" in target
    assert "dupPage" not in target, "the bulk action is reading the page"
    # And the merge itself is handed that same whole-list array.
    assert "mergeHighConfidence(highConfidenceTargets)" in src


def test_the_page_number_is_clamped_to_the_list():
    """Clearing decisions or a re-scan can shorten the list under a page number
    that was valid a moment ago, and an out-of-range page renders as an empty
    panel that looks like a failed load."""
    if not _PAGE.exists():
        return
    src = _PAGE.read_text(encoding="utf-8")
    assert "const dupPageSafe = Math.min(dupPage, dupPageCount - 1);" in src
    assert "slice(dupPageSafe * DUPE_PAGE_SIZE" in src


def test_a_fresh_scan_returns_to_the_first_page():
    if not _PAGE.exists():
        return
    body = _PAGE.read_text(encoding="utf-8").split("const loadDupes = async")[1][:900]
    assert "setDupPage(0);" in body
    assert "maxClusters = DUPE_LIMIT" in body


def test_the_merge_confirmation_says_it_covers_every_group_not_the_page():
    """Once there is a page on screen, "Merge all" is ambiguous in a way it was
    not before — and it is about to discard values on hundreds of records."""
    if not _PAGE.exists():
        return
    src = _PAGE.read_text(encoding="utf-8")
    assert "not just the page on screen" in src
    assert "keeps ONE value per column" in src


def test_the_truncation_banner_survives_for_the_case_it_is_still_true_in():
    """Every group is requested now, so it only fires past 2,000 — but a
    truncated review list that reads as a complete one is exactly what this panel
    was corrected for, and that is still true at 2,001."""
    if not _PAGE.exists():
        return
    src = _PAGE.read_text(encoding="utf-8")
    assert "(dupes.hidden_count ?? 0) > 0" in src
    assert "are not\n              listed" in src or "are not listed" in src


def test_the_download_button_is_wired_to_the_export():
    if not _PAGE.exists() or not _API.exists():
        return
    assert "duplicateSuspectsExport" in _PAGE.read_text(encoding="utf-8")
    assert "duplicate-suspects-export" in _API.read_text(encoding="utf-8")
