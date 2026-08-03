"""The run report — what the tool did to the input file.

It sits beside the bundle download and answers the question that follows one:
the analyst has the FBDI files, and now has to be able to say how they were
produced. Which columns were mapped and on whose authority, what was cleansed,
what merged away as duplicate, what validation found, and which required fields
are still short.

Two claims are worth testing and both are about honesty rather than formatting:

1. **A section that found nothing still gets a sheet, saying so.** An absent
   sheet and a section that ran and found nothing look identical otherwise, and
   that ambiguity is exactly how "the tool never ran that check" gets read as
   "there was nothing to find".
2. **The report describes the run, not a recomputation.** It reads the
   ``dq_report`` persisted on the artifact and the mapping rows as they stand.
   Recomputing cleansing or validation at report time would produce a document
   that disagrees with the files it describes the moment anything has moved.

Pure: builds workbooks in memory from hand-built dicts, and reads the collector's
source for the seam. No database.
"""
import io
import os
from pathlib import Path

import openpyxl

from app.services import conversion_report_service as crs

_BACKEND = Path(__file__).resolve().parent.parent
_APP = _BACKEND / "app"


def _book(**kw):
    data = crs.build_workbook(title="NextPower — what the tool did", **kw)
    return openpyxl.load_workbook(io.BytesIO(data))


def _cells(ws):
    return [str(c.value) for row in ws.iter_rows() for c in row if c.value is not None]


# ── The shape of the document ────────────────────────────────────────────────

def test_it_is_a_real_workbook_with_every_section():
    wb = _book()
    assert wb.sheetnames == ["Summary", "Mappings", "Cleansing", "Duplicates",
                             "Validation", "Required fields", "Run log"]


def test_a_section_that_found_nothing_still_says_so():
    """An absent sheet and a section that ran and found nothing look identical."""
    wb = _book()
    for name, needle in (("Cleansing", "Nothing needed cleansing"),
                         ("Validation", "every checked value passed"),
                         ("Required fields", "Every curated required field"),
                         ("Duplicates", "No generated output"),
                         ("Run log", "Nothing has been generated")):
        assert any(needle in c for c in _cells(wb[name])), f"{name} says nothing"


def test_the_summary_explains_how_to_read_it():
    """The numbers are only useful if a reader knows what they are counting."""
    text = " ".join(_cells(_book()["Summary"]))
    for phrase in ("Source rows", "Merged / de-duplicated", "Columns mapped",
                   "Cleansed values", "Validation errors", "Required short"):
        assert phrase in text


def test_the_summary_says_it_describes_the_run():
    text = " ".join(_cells(_book()["Summary"]))
    assert "not from a fresh recalculation" in text


# ── The content ──────────────────────────────────────────────────────────────

SUMMARY = [{"object": "Supplier", "source_rows": 8561, "output_rows": 8204,
            "merged_or_deduped": 357, "mapped": 96, "total_fields": 104,
            "cleansing_fix_count": 1422, "error_count": 3, "warning_count": 9,
            "required_failed": 1, "blocked": True,
            "output_file": "01_Supplier.zip"}]


def test_the_summary_carries_every_object_and_its_numbers():
    ws = _book(objects=SUMMARY)["Summary"]
    text = " ".join(_cells(ws))
    assert "Supplier" in text and "8561" in text and "8204" in text
    assert "96 of 104" in text


def test_a_blocked_object_is_marked_rather_than_just_listed():
    ws = _book(objects=SUMMARY)["Summary"]
    row = [c for c in ws[5]]
    assert row[0].fill.fgColor.rgb.endswith("FEF2F2"), "a blocked object reads as ordinary"


def test_a_mapping_names_what_fed_it_and_who_decided():
    ws = _book(mappings=[{
        "object": "Supplier", "sheet": "POZ_SUPPLIERS_INT",
        "target_field": "Supplier Name", "source_column": "vendor_name",
        "layer": "workbook", "layer_label": "Mapping workbook",
        "authority": "NXT Supplier Mapping 3 (31Jul26)", "status": "approved",
        "confidence": 1.0, "required": True}])["Mappings"]
    text = " ".join(_cells(ws))
    assert "Supplier Name" in text and "vendor_name" in text
    assert "Mapping workbook" in text
    assert "NXT Supplier Mapping 3 (31Jul26)" in text


def test_a_required_field_nothing_mapped_is_marked_red():
    ws = _book(mappings=[{"object": "Supplier", "target_field": "Supplier Number",
                          "layer": "unmapped", "layer_label": "Not mapped",
                          "required": True}])["Mappings"]
    assert ws.cell(row=5, column=1).fill.fgColor.rgb.endswith("FEF2F2")


def test_a_field_the_model_decided_is_marked_differently_from_one_a_person_did():
    """"How much of this file did the model decide?" is the question a reviewer
    needs answered before sign-off."""
    ws = _book(mappings=[
        {"object": "S", "target_field": "A", "layer": "ai", "layer_label": "AI suggestion"},
        {"object": "S", "target_field": "B", "layer": "manual", "layer_label": "Analyst"},
    ])["Mappings"]
    assert ws.cell(row=5, column=1).fill.fgColor.rgb.endswith("FFFBEB")
    assert not ws.cell(row=6, column=1).fill.fgColor.rgb.endswith("FFFBEB")


def test_cleansing_shows_a_real_before_and_after():
    """A count on its own is not reviewable. "1,422 values changed" could be
    right or catastrophic; "  ACME  -> ACME" can be judged."""
    ws = _book(cleansing=[{"object": "Supplier", "field": "Supplier Name",
                           "rule": "whitespace_punct",
                           "label": "Whitespace & edge punctuation", "count": 1422,
                           "before": "  ACME CORP.  ", "after": "ACME CORP."}])["Cleansing"]
    text = " ".join(_cells(ws))
    assert "  ACME CORP.  " in text and "ACME CORP." in text
    assert "Whitespace & edge punctuation" in text


def test_cleansing_names_the_two_rules_that_rewrite_business_values():
    text = " ".join(_cells(_book()["Cleansing"]))
    assert "Case and legal-suffix" in text and "off unless" in text


def test_duplicates_report_what_went_in_against_what_came_out():
    ws = _book(duplicates=[{"object": "Supplier", "sources": "AllVendors.xlsx",
                            "source_rows": 8561, "output_rows": 8204,
                            "merged_or_deduped": 357, "key": "SupplierNumber",
                            "decisions": 12}])["Duplicates"]
    text = " ".join(_cells(ws))
    assert "8561" in text and "8204" in text and "357" in text
    assert "SupplierNumber" in text


def test_an_unknown_duplicate_count_is_a_dash_not_a_zero():
    """Zero merged and "the run cannot account for the difference" are different
    facts, and printing one as the other is a lie the reader cannot detect."""
    ws = _book(duplicates=[{"object": "Supplier", "merged_or_deduped": None,
                            "source_rows": None, "output_rows": None}])["Duplicates"]
    assert "—" in _cells(ws)


def test_validation_says_what_it_means_and_what_to_do():
    ws = _book(validation=[{"object": "Supplier", "severity": "error",
                            "issue_type": "Missing Required Field",
                            "field_name": "Supplier Number", "impacted_count": 41,
                            "message": "Blank on 41 rows.",
                            "suggested_fix": "Map it or set a default."}])["Validation"]
    text = " ".join(_cells(ws))
    assert "Missing Required Field" in text and "Blank on 41 rows." in text
    assert "Map it or set a default." in text
    assert ws.cell(row=5, column=1).fill.fgColor.rgb.endswith("FEF2F2")


def test_validation_says_it_is_advisory():
    text = " ".join(_cells(_book()["Validation"]))
    assert "does not stop the download" in text


def test_a_not_owned_required_field_explains_why_nothing_was_checked():
    ws = _book(required=[{"object": "Customer", "sheet": "HZ_IMP_CONTACTS_T",
                          "field": "Role Type", "status": "not owned",
                          "detail": "Curated for a sheet this conversion does not "
                                    "produce, so nothing was checked."}])["Required fields"]
    assert "nothing was checked" in " ".join(_cells(ws))


def test_a_generation_that_shipped_without_a_step_says_so_in_the_run_log():
    """A file generated WITHOUT the learning library is not the same file, and
    the reason used to live only in a server log nobody reads."""
    ws = _book(run_log=[{"object": "Supplier", "output_file": "01_Supplier.zip",
                         "state": "generated", "rows": 8204,
                         "problem": "RuntimeError: apply_learned_to_conversion failed"}])["Run log"]
    text = " ".join(_cells(ws))
    assert "apply_learned_to_conversion failed" in text
    assert ws.cell(row=5, column=1).fill.fgColor.rgb.endswith("FEF2F2")


def test_the_run_log_says_a_recorded_failure_still_shipped():
    assert "shipped anyway" in " ".join(_cells(_book()["Run log"]))


def test_a_stale_output_is_flagged():
    ws = _book(run_log=[{"object": "Supplier", "state": "stale"}])["Run log"]
    assert ws.cell(row=5, column=1).fill.fgColor.rgb.endswith("FFFBEB")


# ── The seam: it reports the run, and it is actually wired up ────────────────

def _ops():
    return (_APP / "routers" / "operations.py").read_text(encoding="utf-8")


def _collector():
    return _ops().split("async def _report_rows_for_conversion(")[1] \
                 .split("\n@output_router")[0]


def test_the_report_reads_the_artifact_rather_than_recomputing():
    """Recomputing would produce a document that disagrees with the files it
    describes the moment anything has moved."""
    body = _collector()
    assert "latest.dq_report" in body
    assert 'dq.get("cleansing_fixes")' in body
    assert 'dq.get("top_issues")' in body
    # The expensive re-scans are deliberately not run here.
    assert "find_duplicate_clusters" not in body
    assert "apply_cleansing" not in body
    assert "validate_frame" not in body


def test_one_object_failing_does_not_lose_the_whole_report():
    body = _ops().split("async def project_conversion_report(")[1].split("\n@")[0]
    assert '"state": "report failed"' in body
    assert "continue" in body


def test_a_failed_section_appears_in_the_report_rather_than_vanishing():
    """A missing section reads as "nothing found"."""
    body = _collector()
    assert "req_problem" in body
    assert "required-field check failed" in body


def test_the_mapping_row_records_the_authority_behind_the_decision():
    body = _collector()
    assert '"derived_from"' in body or "derived_from" in body
    assert "approved_by" in body


def test_the_endpoint_streams_a_spreadsheet_with_a_real_filename():
    body = _ops().split("async def project_conversion_report(")[1].split("\n@")[0]
    assert "spreadsheetml.sheet" in body
    assert "Content-Disposition" in body
    assert "_safe_name(" in body


def test_the_build_runs_off_the_event_loop():
    """A 19-sheet project builds thousands of rows; doing that on the loop is
    what makes a request hang."""
    body = _ops().split("async def project_conversion_report(")[1].split("\n@")[0]
    assert "asyncio.to_thread" in body


def test_objects_are_reported_in_load_order():
    body = _ops().split("async def project_conversion_report(")[1].split("\n@")[0]
    assert "planned_load_order" in body


def test_the_button_exists_and_calls_the_endpoint():
    """A capability with no caller is an inert feature — this repo's habit."""
    page = (_BACKEND.parent / "frontend" / "src" / "pages"
            / "ProjectOverviewPage.tsx").read_text(encoding="utf-8")
    assert "Output report (.xlsx)" in page
    assert "onClick={downloadReport}" in page
    api = (_BACKEND.parent / "frontend" / "src" / "api"
           / "index.ts").read_text(encoding="utf-8")
    assert "conversionReport:" in api
    assert "/conversion-report" in api


def test_the_button_sits_with_the_bundle_downloads():
    """It answers the question the bundle raises, so it belongs beside it rather
    than on a page nobody opens after a download."""
    page = (_BACKEND.parent / "frontend" / "src" / "pages"
            / "ProjectOverviewPage.tsx").read_text(encoding="utf-8")
    # Both places the bundle can be pulled from carry it.
    assert page.count("onClick={downloadReport}") == 2
    # In the object-list header it sits between the two bundle buttons.
    group = page[page.index("Generate all & download (.zip)"):]
    assert group.index("Output report (.xlsx)") < group.index("templates (.zip)")


def test_the_report_does_not_silently_regenerate_the_bundle():
    """Pressing it after a download must describe the files you just got, not
    build different ones and describe those."""
    body = _ops().split("async def project_conversion_report(")[1].split("\n@")[0]
    assert "generate_output_artifact" not in body
    assert "generate_merged" not in body
