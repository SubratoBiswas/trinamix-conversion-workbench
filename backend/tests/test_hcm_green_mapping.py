""""The HCM (Employee) mapping is what the workbook COLOURED green, not what it listed.

`NXT HCM Field Mapping 1.xlsx`, tab 'Source Files Mapping', carries its own legend:

    Green  = Mapped                       <- the only rows we import
    Yellow = Questions to NEXTPOWER
    Orange = Duplicate
    Blue   = Oracle required, missing from source
    Red    = Not to Bring

47 data rows, 20 of them green. Green is read by CELL FILL (FF92D050) on the Source
Column Name cell — reading it by eye off a screenshot is how a Red row gets loaded.

Five things in the sheet are not a plain column mapping, and every one of them has a
matching failure mode already on this project's record:

  * The mapping tab writes source columns with UNDERSCORES (``Employee_ID``); the
    Workday extract header row uses SPACES (``Employee ID``). The apply path
    normalises, so a test that only exercised apply would pass — but
    ``propagate_learning_to_open_conversions`` writes ``original_value`` VERBATIM onto
    the mapping row, so the workbook spelling would put a source column on screen
    that the dataset does not contain.
  * ONE source, TWO targets: ``Country -> Country, LegislationCode``.
  * A CONSTANT where a source column goes: ``default value ( 1/1/1900 )``. Seeding it
    as a column would hunt the extract for a column of that name.
  * The ``(Object)`` suffix is a SCOPE, not a label. ``EffectiveStartDate`` exists on
    nine components; ``EffectiveStartDate(Location)`` means those two rows and no
    others, or every worker's record gets date-stamped 1900.
  * "Sheets", for HDL in this tool, are COMPONENT names — ``hdl_seed_service`` writes
    one FBDISheet per component. ``FirstName`` is on PersonName, not on Worker. The
    first cut of this file scoped every learning to the top-level ``.dat`` object,
    which excluded eleven of nineteen learnings from the component their field
    actually lives on. ``test_every_seeded_field_exists_on_the_component_it_claims``
    is the test that caught it, and it is the reason this file cross-checks against
    ``hdl_schema`` rather than against itself.

That last one is the recurring shape here: a capability ships, passes its tests
against hand-made inputs, and never meets the real schema.
"""
import json
import os
import sys
from pathlib import Path

sys.path.insert(0, os.path.join(os.path.dirname(__file__), ".."))

from app.services.hdl_schema import (  # noqa: E402
    HDL_LOAD_ORDER, JOB, LOCATION, PERSON_NAME, WORK_RELATIONSHIP, all_components,
)
from app.services.hdl_output_service import render_cell                # noqa: E402
from app.services.learning_service import sheet_allowed                # noqa: E402

_ROOT = Path(__file__).resolve().parent.parent
_DOC = json.loads((_ROOT / "app" / "data" / "hcm_source_mapping.json").read_text(encoding="utf-8"))
_MAPS = _DOC["mappings"]
_CONSTS = _DOC["constants"]

# Measured off the workbook: rows whose Source Column Name cell is filled FF92D050,
# excluding the legend swatch on row 2. 20 rows; every one also reads
# "Bring to Oracle = Yes"; the header row is row 8.
GREEN_ROWS = {9, 12, 13, 14, 17, 18, 19, 22, 24, 34, 37, 38, 39, 40, 43, 44, 45, 46, 54, 55}

# Row 3 of the 'Workday Extract' tab — the real header spellings, all 31 of them.
EXTRACT_HEADERS = {
    "Employee ID", "Preferred First Name", "Preferred Last Name", "Preferred Name",
    "Legal First Name", "Legal Last Name", "Legal Middle Name", "Legal Name", "Company",
    "Hire Date", "Worker Type", "Employee Type", "Manager Name", "Active Status",
    "On Leave", "Manager - Level 01", "Manager - Level 02", "Manager - Level 03",
    "Manager - Level 04", "Manager - Level 05", "Manager - Level 06", "Manager - Level 07",
    "Manager - Level 08", "Manager - Level 09", "Location", "Position", "Business Title",
    "Country", "Email", "Cost Center", "Cost Center Name",
}

# field name -> {component names carrying it}, and component -> owning .dat object.
FIELD_COMPONENTS: dict[str, set] = {}
COMPONENT_OBJECT: dict[str, str] = {}
for _obj, _comp, _fields in all_components():
    COMPONENT_OBJECT[_comp] = _obj
    for _f in _fields:
        FIELD_COMPONENTS.setdefault(_f["name"], set()).add(_comp)


def check(name, cond, detail=""):
    if cond:
        print(f"  PASS  {name}"); return
    raise AssertionError(f"{name} {detail}".strip())


class L:
    """The two attributes ``sheet_allowed`` reads."""
    def __init__(self, sheets=None, exclude_sheets=None):
        self.sheets = sheets or []
        self.exclude_sheets = exclude_sheets or []


def spec(component, field_name):
    _, fields = component
    for f in fields:
        if f["name"] == field_name:
            return f
    return None


def targets_of(col):
    return sorted(m["target_field"] for m in _MAPS if m["source_column"] == col)


def test_only_the_green_rows_came_across():
    """20 of 47. A count that drifts means somebody read colour off a screenshot."""
    rows = {m["source_row"] for m in _MAPS} | {c["source_row"] for c in _CONSTS}
    check("every green row is represented", rows == GREEN_ROWS,
          f"missing {sorted(GREEN_ROWS - rows)}, extra {sorted(rows - GREEN_ROWS)}")
    check("20 green rows", len(GREEN_ROWS) == 20)
    check("no row was imported twice as both a mapping and a constant",
          not ({m["source_row"] for m in _MAPS} & {c["source_row"] for c in _CONSTS}))


def test_source_columns_are_spelled_the_way_the_extract_spells_them():
    """Underscores are the mapping tab's convention, not the data's."""
    for m in _MAPS:
        if not m.get("verified_against_extract"):
            continue
        col = m["source_column"]
        check(f"{col!r} is a real extract header", col in EXTRACT_HEADERS)
        check(f"{col!r} carries no underscore", "_" not in col)
    aliased = [m for m in _MAPS if m.get("verified_against_extract")
               and "_" in (m.get("as_written_in_workbook") or "")]
    check("the workbook's own spelling is still recorded as an alias",
          len(aliased) >= 8, f"only {len(aliased)}")


def test_the_columns_we_could_not_verify_are_flagged_rather_than_assumed():
    """Five green rows name columns absent from THIS extract — their Source File cell
    is blank in the workbook because they come from other Workday files."""
    unverified = sorted(m["source_column"] for m in _MAPS if not m.get("verified_against_extract"))
    check("exactly the five off-extract columns",
          unverified == ["Job Profile", "Job_Code", "Location_Code", "Manager_ID", "Military_Service"],
          f"got {unverified}")
    for col in unverified:
        check(f"{col!r} is genuinely not in this extract", col not in EXTRACT_HEADERS)


def test_every_seeded_field_exists_on_the_component_it_claims():
    """THE seam. A learning scoped to a component that does not carry the field is
    excluded from the one sheet it was written for, and looks fine in the list.

    Scoping by the ``.dat`` object instead of the component did exactly that to
    eleven of nineteen rows: FirstName is on PersonName, EmailAddress on PersonEmail,
    LegalEmployerName on WorkRelationship/WorkTerms/Assignment — none of them on a
    sheet named "Worker".
    """
    for entry in _MAPS + _CONSTS:
        fld = entry["target_field"]
        comps = entry.get("hdl_components") or []
        if entry.get("oracle_field_exists") is False:
            check(f"{fld} has no Oracle field and claims no component", comps == [])
            check(f"{fld} really is absent from every component", fld not in FIELD_COMPONENTS)
            continue
        check(f"{fld} claims at least one component", bool(comps))
        for c in comps:
            check(f"{c} is a real HDL component", c in COMPONENT_OBJECT, f"got {c!r}")
            check(f"{c} carries {fld}", c in FIELD_COMPONENTS.get(fld, set()),
                  f"{fld} lives on {sorted(FIELD_COMPONENTS.get(fld, set()))}")
        obj = entry.get("hdl_object")
        if obj:
            check(f"{fld}'s .dat file matches its components",
                  all(COMPONENT_OBJECT[c] == obj for c in comps)
                  or len({COMPONENT_OBJECT[c] for c in comps}) > 1,
                  f"{obj} vs {[COMPONENT_OBJECT[c] for c in comps]}")


def test_the_object_suffix_narrows_rather_than_labels():
    """``EffectiveStartDate(Location)`` means Location and nothing else."""
    check("EffectiveStartDate is on nine components",
          len(FIELD_COMPONENTS["EffectiveStartDate"]) == 9,
          f"got {sorted(FIELD_COMPONENTS['EffectiveStartDate'])}")
    scoped = sorted(c["hdl_components"][0] for c in _CONSTS)
    check("but the 1900 constant is scoped to two", scoped == ["Job", "Location"], f"got {scoped}")
    # Same for the fields the workbook suffixed: each is on several components.
    for fld, want in (("Name", "Job"), ("JobCode", "Job"), ("LocationCode", "Location")):
        got = [m for m in _MAPS if m["target_field"] == fld]
        check(f"{fld} imported once", len(got) == 1)
        check(f"{fld} is on more than one component", len(FIELD_COMPONENTS[fld]) > 1)
        check(f"{fld} is scoped to {want} alone", got[0]["hdl_components"] == [want],
              f"got {got[0]['hdl_components']}")


def test_country_became_two_learnings_not_one():
    """``Country -> Country, LegislationCode`` is one cell holding two targets — and
    they land on different .dat files, which is why one learning could not serve both."""
    check("both targets survived", targets_of("Country") == ["Country", "LegislationCode"],
          f"got {targets_of('Country')}")
    rows = {m["source_row"] for m in _MAPS if m["source_column"] == "Country"}
    check("both point back at the same workbook row", rows == {37}, f"got {rows}")
    objs = {m["target_field"]: m["hdl_object"] for m in _MAPS if m["source_column"] == "Country"}
    check("Country goes to Location.dat", objs["Country"] == "Location")
    check("LegislationCode to Worker.dat", objs["LegislationCode"] == "Worker")


def test_the_default_date_rows_are_constants_and_not_column_mappings():
    for m in _MAPS:
        check("no mapping treats the default-date text as a column",
              "default value" not in m["source_column"].lower(), f"got {m['source_column']!r}")
    pairs = sorted((c["hdl_object"], c["target_field"], c["value"]) for c in _CONSTS)
    check("both EffectiveStartDate defaults are present",
          pairs == [("Job", "EffectiveStartDate", "1900/01/01"),
                    ("Location", "EffectiveStartDate", "1900/01/01")], f"got {pairs}")


def test_the_loader_actually_writes_the_1900_default():
    """The JSON is documentation until the generator agrees with it. Location and Job
    are deduped setup objects — the hire date they used to carry came from whichever
    employee row survived the dedupe, so it changed with the extract."""
    for comp, label in ((LOCATION, "Location"), (JOB, "Job")):
        f = spec(comp, "EffectiveStartDate")
        check(f"{label}.EffectiveStartDate is a constant", f["kind"] == "const", f"got {f['kind']}")
        check(f"{label} renders 1900/01/01 whatever the row says",
              render_cell(f, lambda n, s: "2017/01/01") == "1900/01/01")


def test_preferred_name_has_somewhere_to_land():
    """Row 12 maps Preferred_Name -> KnownAs, and PersonName had no KnownAs."""
    f = spec(PERSON_NAME, "KnownAs")
    check("PersonName now carries KnownAs", f is not None)
    check("sourced from the extract's Preferred Name", f["source"] == "Preferred Name")
    check("and it renders", render_cell(f, lambda n, s: "Antonio Sendín") == "Antonio Sendín")


def test_military_service_is_sourced_rather_than_hard_blank():
    """Row 54. It was ``blank``, so a supplied value would have been discarded — and
    it renders empty today only because the column is in another Workday file."""
    f = spec(WORK_RELATIONSHIP, "OnMilitaryServiceFlag")
    check("no longer hard-blank", f["kind"] == "source", f"got {f['kind']}")
    check("reads Military_Service", f["source"] == "Military_Service")
    check("a supplied value survives", render_cell(f, lambda n, s: "Y") == "Y")
    check("an absent column still renders empty", render_cell(f, lambda n, s: None) == "")


def test_a_green_row_with_no_oracle_field_is_recorded_but_not_seeded():
    """Cost Center. The workbook's own comment: no suitable Oracle field exists. It
    belongs in the Learning Center as intent, not as a mapping onto a phantom field."""
    cc = [m for m in _MAPS if m["source_column"] == "Cost Center"]
    check("the row is kept", len(cc) == 1)
    check("flagged as having no Oracle field", cc[0].get("oracle_field_exists") is False)
    seed = (_ROOT / "app" / "services" / "catalog_seed_service.py").read_text(encoding="utf-8")
    check("and the seeder skips it",
          'if m.get("oracle_field_exists") is False:' in seed and "no_field += 1" in seed)


def test_workrelationship_is_a_block_of_worker_not_a_sixth_dat_file():
    check("five .dat files", HDL_LOAD_ORDER == ["Location", "Job", "Position",
                                                "PositionHierarchy", "Worker"])
    check("WorkRelationship is not one of them", "WorkRelationship" not in HDL_LOAD_ORDER)
    check("it is a component of Worker.dat", COMPONENT_OBJECT["WorkRelationship"] == "Worker")
    mil = [m for m in _MAPS if m["source_column"] == "Military_Service"][0]
    check("the green row lands in Worker.dat", mil["hdl_object"] == "Worker")
    check("on the WorkRelationship block", mil["hdl_components"] == ["WorkRelationship"])


def test_component_scope_holds_through_the_real_sheet_allowed():
    """Not the mechanism in isolation — the seeded scope through the live function."""
    first = [m for m in _MAPS if m["target_field"] == "FirstName"][0]
    lm = L(sheets=first["hdl_components"])
    check("FirstName applies to PersonName", sheet_allowed(lm, "PersonName") is True)
    check("and not to Worker", sheet_allowed(lm, "Worker") is False)
    check("and not to Assignment", sheet_allowed(lm, "Assignment") is False)
    check("an unknown sheet does not silently gain it", sheet_allowed(lm, None) is False)
    multi = L(sheets=[m for m in _MAPS if m["target_field"] == "LegalEmployerName"][0]["hdl_components"])
    for s in ("WorkRelationship", "WorkTerms", "Assignment"):
        check(f"LegalEmployerName applies to {s}", sheet_allowed(multi, s) is True)
    check("but not to PersonName", sheet_allowed(multi, "PersonName") is False)


def test_the_seeder_and_its_wiring_exist():
    """Data with no seeder is an inert feature one layer up — this repo's habit."""
    seed = (_ROOT / "app" / "services" / "catalog_seed_service.py").read_text(encoding="utf-8")
    check("seeder exists", "async def seed_hcm_source_mapping(" in seed)
    check("it scopes per COMPONENT", '"sheets": list(comps or [])' in seed)
    check("it tags the source system", '"source_erp": "workday"' in seed)
    check("constants are seeded as defaults, not columns", '"example_default"' in seed)
    check("the two 1900 constants stay two rows", 'f"(constant:{comps[0]})"' in seed)
    check("it honours tombstones", "include_deleted=True" in seed)
    main = (_ROOT / "app" / "main.py").read_text(encoding="utf-8")
    check("it runs at startup", "seed_hcm_source_mapping()" in main)
    api = (_ROOT / "app" / "routers" / "learned.py").read_text(encoding="utf-8")
    check("and on demand", "reseed-hcm-mapping" in api)


def test_the_workbook_still_says_what_we_recorded():
    """Re-derive green from the workbook when it is to hand. Skips when it is not —
    the vendored GREEN_ROWS above is the standing assertion; this one catches a
    workbook that changed under us."""
    up = Path("/root/.claude/uploads")
    candidates = sorted(up.rglob("*HCM_Field_Mapping*.xlsx")) if up.exists() else []
    if not candidates:
        print("  SKIP  workbook not present in this environment")
        return
    import openpyxl
    ws = openpyxl.load_workbook(candidates[0], data_only=True)["Source Files Mapping"]
    green, bring = set(), []
    for r in range(9, ws.max_row + 1):
        cell = ws.cell(r, 3)                       # Source Column Name
        if getattr(cell.fill.start_color, "rgb", None) == "FF92D050":
            green.add(r)
            bring.append(str(ws.cell(r, 5).value or "").strip().lower())
    check("the workbook still colours exactly these rows", green == GREEN_ROWS,
          f"workbook {sorted(green)}")
    check("every green row also reads Bring to Oracle = Yes", set(bring) == {"yes"}, f"got {set(bring)}")


if __name__ == "__main__":
    for fn in [v for k, v in sorted(globals().items()) if k.startswith("test_")]:
        print(fn.__name__); fn()
    print("\nall HCM green-mapping checks passed")
