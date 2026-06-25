"""Robust Oracle FBDI template parser supporting two formats:

1. Oracle FBDI transposed format: column A has row-labels (Name/Description/Data Type),
   fields span columns B onwards.
2. Standard tabular format: row 1 has field names as column headers from col A onwards,
   prefix "* " marks required fields.
"""
from __future__ import annotations
import re
from pathlib import Path
from openpyxl import load_workbook

_DATA_TYPE_RE = re.compile(r"^\s*([A-Za-z]+)\s*(?:\(\s*(\d+)(?:\s*,\s*\d+)?\s*\))?", re.IGNORECASE)


def _parse_data_type(raw):
    if not raw:
        return ("Character", None)
    m = _DATA_TYPE_RE.match(str(raw))
    if not m:
        return (str(raw).strip(), None)
    dtype = m.group(1).strip().capitalize()
    length = int(m.group(2)) if m.group(2) else None
    return (dtype, length)


def _looks_like_module_label(value):
    if not value:
        return False
    s = str(value).strip()
    keywords = ("management", "planning", "order promising", "operations", "supply", "common", "interface", "loader")
    return any(k in s.lower() for k in keywords) and len(s) < 80


def _read_sheet_rows(ws):
    """Read all rows from a worksheet using iter_rows (avoids max_row/max_col None issues)."""
    rows = []
    for row in ws.iter_rows(values_only=True):
        rows.append(row)
    # Strip trailing all-None rows
    while rows and all(v is None for v in rows[-1]):
        rows.pop()
    return rows


def parse_fbdi_template(file_path):
    file_path = Path(file_path)
    wb = load_workbook(filename=file_path, data_only=True, keep_vba=False, read_only=False)

    business_object = None
    description = None

    for sname in wb.sheetnames:
        if "instruction" in sname.lower():
            ws = wb[sname]
            for row in ws.iter_rows(min_row=1, max_row=10, values_only=True):
                for cell in row:
                    if cell and isinstance(cell, str) and ":" in cell and len(cell) < 200:
                        if any(k in cell.lower() for k in ("upload:", "import:", "interface:")):
                            business_object = cell.split(":", 1)[1].strip()
                            break
                if business_object:
                    break
            break

    sheets_out = []
    fields_out = []
    seq = 0

    for sname in wb.sheetnames:
        if "instruction" in sname.lower():
            continue
        ws = wb[sname]

        # Use iter_rows to avoid max_row/max_col returning None on some openpyxl versions
        all_rows = _read_sheet_rows(ws)
        if not all_rows:
            continue
        max_row = len(all_rows)
        max_col = max((len(r) for r in all_rows), default=0)
        if max_col == 0:
            continue

        def cell_val(r, c):
            """1-based row/col access into all_rows list."""
            try:
                return all_rows[r - 1][c - 1]
            except IndexError:
                return None

        # Detect format: Oracle transposed has "Name" as a row label in col A
        col_a_labels = []
        for r in range(1, min(15, max_row + 1)):
            v = cell_val(r, 1)
            if v:
                col_a_labels.append((r, str(v).strip()))

        is_oracle_fbdi = any(lbl.lower() == "name" for _, lbl in col_a_labels)
        sheet_field_count = 0

        if is_oracle_fbdi:
            name_row = next((r for r, lbl in col_a_labels if lbl.lower() == "name"), 1)
            desc_row = next((r for r, lbl in col_a_labels if "description" in lbl.lower()), 2)
            type_row = next((r for r, lbl in col_a_labels if "data type" in lbl.lower() or lbl.lower() == "type"), 3)
            module_rows = [(r, lbl) for r, lbl in col_a_labels if r > type_row and _looks_like_module_label(lbl)]

            for col in range(2, max_col + 1):
                raw_name = cell_val(name_row, col)
                if not raw_name:
                    continue
                name = str(raw_name).strip()
                if not name:
                    continue
                is_req = name.startswith("*")
                field_name = name.lstrip("*").strip()
                desc_val = cell_val(desc_row, col)
                desc_text = str(desc_val).strip() if desc_val else None
                dtype_raw = cell_val(type_row, col)
                data_type, max_length = _parse_data_type(dtype_raw)
                req_modules = []
                for r, lbl in module_rows:
                    cv = cell_val(r, col)
                    if cv and "required" in str(cv).strip().lower():
                        req_modules.append(lbl)
                required = is_req or bool(req_modules)
                seq += 1
                sheet_field_count += 1
                fields_out.append({
                    "field_name": field_name,
                    "display_name": field_name,
                    "description": desc_text,
                    "required": required,
                    "data_type": data_type,
                    "max_length": max_length,
                    "format_mask": "YYYYMMDD" if data_type.lower() == "date" else None,
                    "sample_value": None,
                    "lookup_type": None,
                    "validation_notes": None,
                    "sequence": seq,
                    "sheet_name": sname,
                    "required_modules": req_modules,
                })

        else:
            # Standard tabular: row 1 = headers, optional row 2 = sample
            header_row = all_rows[0] if all_rows else []
            sample_row = all_rows[1] if len(all_rows) > 1 else []

            for col_idx, raw_name in enumerate(header_row):
                if not raw_name:
                    continue
                name = str(raw_name).strip()
                if not name:
                    continue
                is_req = name.startswith("*")
                field_name = name.lstrip("* ").strip()
                if not field_name:
                    continue
                sample_val = sample_row[col_idx] if col_idx < len(sample_row) else None
                sample_text = str(sample_val).strip() if sample_val is not None else None
                if isinstance(sample_val, (int, float)):
                    data_type, max_length = "Number", None
                else:
                    data_type, max_length = "Character", None
                seq += 1
                sheet_field_count += 1
                fields_out.append({
                    "field_name": field_name,
                    "display_name": field_name,
                    "description": None,
                    "required": is_req,
                    "data_type": data_type,
                    "max_length": max_length,
                    "format_mask": None,
                    "sample_value": sample_text,
                    "lookup_type": None,
                    "validation_notes": None,
                    "sequence": seq,
                    "sheet_name": sname,
                    "required_modules": [],
                })

        sheets_out.append({"sheet_name": sname, "sequence": len(sheets_out), "field_count": sheet_field_count})

    return {"business_object": business_object, "description": description, "sheets": sheets_out, "fields": fields_out}
