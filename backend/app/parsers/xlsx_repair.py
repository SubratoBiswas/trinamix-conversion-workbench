"""Open a real-world Oracle .xlsm that openpyxl refuses.

WHY
---
``SupplierSiteImportTemplate.xlsm``, as downloaded from Oracle, contains one
``<font><family val="34"/></font>``. OOXML bounds that attribute at 14, so openpyxl
raises and refuses the ENTIRE workbook:

    ValueError: Unable to read workbook: could not read stylesheet ...
    ValueError: Max value is 14

Excel opens it without complaint. The consequence in the tool was total: uploading
that template — a file the analyst actually has — failed outright, and the parse error
said "invalid XML", which points the reader at the wrong thing.

There is already a ``_repair_xlsx`` in tabular_parser that swaps in a minimal
stylesheet. That is fine for reading a data extract and wrong for a TEMPLATE: the
template's own formatting is the thing ``template_fill_service`` fills and hands back,
so discarding it would change the deliverable. So repair in order of least damage:

  1. As-is.
  2. Clamp only the out-of-range attribute values, keeping every other style.
  3. Only then fall back to a minimal stylesheet (metadata reads survive; formatting
     does not, so the caller is told).

Pure: stdlib + openpyxl.
"""
from __future__ import annotations

import io
import logging
import re
import zipfile
from typing import Any

from openpyxl import load_workbook

log = logging.getLogger(__name__)

# OOXML CT_FontFamily is 0..14; CT_FontCharSet is a byte. Excel writes values outside
# both. Clamp rather than delete so the font still resolves to something sane.
_FAMILY = re.compile(rb'(<family\s+val=")(\d+)(")')
_CHARSET = re.compile(rb'(<charset\s+val=")(-?\d+)(")')

_MIN_STYLES = (
    b'<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
    b'<styleSheet xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main">'
    b'<fonts count="1"><font><sz val="11"/><name val="Calibri"/></font></fonts>'
    b'<fills count="1"><fill><patternFill patternType="none"/></fill></fills>'
    b'<borders count="1"><border/></borders>'
    b'<cellStyleXfs count="1"><xf numFmtId="0" fontId="0" fillId="0" borderId="0"/></cellStyleXfs>'
    b'<cellXfs count="1"><xf numFmtId="0" fontId="0" fillId="0" borderId="0" xfId="0"/></cellXfs>'
    b'</styleSheet>'
)


def clamp_style_values(styles_xml: bytes) -> tuple[bytes, int]:
    """Bring out-of-range font attributes inside the schema. ``(xml, n_changed)``."""
    changed = 0

    def _fam(m):
        nonlocal changed
        v = int(m.group(2))
        if 0 <= v <= 14:
            return m.group(0)
        changed += 1
        return m.group(1) + b"2" + m.group(3)      # 2 = Swiss, a safe default

    def _chs(m):
        nonlocal changed
        v = int(m.group(2))
        if 0 <= v <= 255:
            return m.group(0)
        changed += 1
        return m.group(1) + b"0" + m.group(3)

    out = _FAMILY.sub(_fam, styles_xml)
    out = _CHARSET.sub(_chs, out)
    return out, changed


# EVERY part that can carry a <font>, not only the stylesheet.
#
# The out-of-range value is a FONT attribute, and OOXML lets a font description
# appear in more places than xl/styles.xml: shared strings carry one per rich-text
# run, and each comment carries one per run of its tooltip. openpyxl parses all of
# them through the same bounded descriptor, so any single one refuses the whole
# workbook.
#
# 3_SupplierSite_POZ_SUPPLIER_SITES_INT.xlsm is the case that proved it. Its
# styles.xml is clean — the <family val="34"/> sits in xl/comments1.xml, in a
# tooltip Oracle attaches to a column header. Repairing only the stylesheet
# reported "nothing out of range", changed nothing, and the workbook still would
# not open, so a filled Supplier Site template could never be produced.
#
# Nothing said so, either, and that is the part worth remembering: the parser
# reads with read_only=True, which skips comments entirely. The template uploaded,
# parsed, and listed all 211 of its fields — and failed only at the moment of
# being filled, which is a different code path on a different day.
_FONT_PARTS = re.compile(r"^xl/(styles\.xml|sharedStrings\.xml|comments\d*\.xml)$")


def _rewrite(raw: bytes, styles: bytes) -> bytes:
    """Swap the stylesheet alone. Kept for the last-resort _MIN_STYLES fallback,
    which is deliberately about the stylesheet and nothing else."""
    return _rewrite_parts(raw, {"xl/styles.xml": styles})


def _rewrite_parts(raw: bytes, replacements: dict[str, bytes]) -> bytes:
    """Copy the zip through, swapping only the named parts.

    Everything else — macros, drawings, the VBA project — passes through byte for
    byte, because this file is a TEMPLATE that gets filled and handed back, and
    its own formatting is the deliverable.
    """
    src = zipfile.ZipFile(io.BytesIO(raw))
    buf = io.BytesIO()
    with zipfile.ZipFile(buf, "w", zipfile.ZIP_DEFLATED) as out:
        for item in src.infolist():
            data = replacements.get(item.filename)
            if data is None:
                data = src.read(item.filename)
            out.writestr(item.filename, data)
    return buf.getvalue()


def repair_bytes(raw: bytes) -> tuple[bytes, str]:
    """``(possibly-rewritten bytes, what was done)``. Never raises."""
    try:
        with zipfile.ZipFile(io.BytesIO(raw)) as z:
            parts = [n for n in z.namelist() if _FONT_PARTS.match(n)]
            if not parts:
                return raw, "no part that could carry a font"
            originals = {n: z.read(n) for n in parts}
    except Exception:                                           # noqa: BLE001
        return raw, "not a readable zip"

    replacements: dict[str, bytes] = {}
    total = 0
    for name, data in originals.items():
        clamped, n = clamp_style_values(data)
        if n:
            replacements[name] = clamped
            total += n
    if total:
        return (_rewrite_parts(raw, replacements),
                f"clamped {total} out-of-range font attribute(s) in "
                f"{', '.join(sorted(replacements))}")
    return raw, "nothing out of range"


def load_workbook_tolerant(source: Any, **kwargs) -> Any:
    """``load_workbook`` that survives Excel's out-of-spec stylesheets.

    ``source`` may be a path or bytes. Escalates only as far as needed, and logs
    which step succeeded so a formatting-sensitive caller can tell whether the
    original styles survived.
    """
    if isinstance(source, (bytes, bytearray)):
        raw = bytes(source)
        label = "<bytes>"
    else:
        label = str(source)
        with open(source, "rb") as fh:
            raw = fh.read()

    try:
        return load_workbook(io.BytesIO(raw), **kwargs)
    except Exception as first:                                  # noqa: BLE001
        repaired, what = repair_bytes(raw)
        if repaired is not raw:
            try:
                wb = load_workbook(io.BytesIO(repaired), **kwargs)
                log.warning("%s: openpyxl refused the workbook (%s); %s and it opened. "
                            "Styles are otherwise untouched.", label, first, what)
                return wb
            except Exception as second:                         # noqa: BLE001
                first = second
        try:
            wb = load_workbook(io.BytesIO(_rewrite(raw, _MIN_STYLES)), **kwargs)
            log.warning("%s: opened only after replacing the stylesheet (%s). Cell "
                        "VALUES are intact; original FORMATTING is not, so do not use "
                        "this copy to fill and hand back a template.", label, first)
            return wb
        except Exception:                                       # noqa: BLE001
            raise first
