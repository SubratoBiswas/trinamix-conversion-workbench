"""CSV / XLSX parsing & column profiling.

Profiles each column with: inferred type, null %, distinct count, sample values,
min/max for numeric/date, and a pattern summary.
"""
from __future__ import annotations

import logging
import re
from pathlib import Path
from typing import Any

import pandas as pd

logger = logging.getLogger(__name__)

# Header keywords, highest priority first, for the column most likely to hold an
# UNescaped delimiter — a free-text field. Used to recover over-length rows (see
# _repair_overlong_delimited) rather than silently dropping them.
_FREETEXT_PRIORITY = ("title", "description", "comment", "note", "remark",
                      "address", "name")


def _freetext_overflow_col(header: list[str]) -> int | None:
    """Index of the free-text column an unescaped delimiter most likely landed in, or
    None when the header has no plausible free-text column (then an over-length row is
    left for on_bad_lines to skip rather than risk corrupting a structured column)."""
    low = [str(h).lower() for h in header]
    for kw in _FREETEXT_PRIORITY:
        for i, h in enumerate(low):
            if kw in h:
                return i
    return None


def _repair_overlong_delimited(text: str, sep: str) -> str:
    """Recover rows a delimited export broke by leaving an unescaped ``sep`` inside a
    free-text field — e.g. a Workday job title "Planner (Modules | Dampers | Motors)"
    in a pipe file. Such a row has MORE fields than the header, so pandas'
    ``on_bad_lines='skip'`` DROPS a real record (observed: 3 employees vanished from a
    2,206-row extract). When the surplus can be attributed to ONE free-text column, the
    extra pieces are merged back into it so the row parses. Rows containing a quote are
    left untouched (pandas handles quoted delimiters); rows that cannot be repaired to
    exactly the header width are left as-is."""
    lines = text.split("\n")
    hi = next((i for i, ln in enumerate(lines) if ln.strip() != ""), None)
    if hi is None:
        return text
    header = lines[hi].split(sep)
    ncol = len(header)
    if ncol < 2:
        return text
    idx = _freetext_overflow_col(header)
    if idx is None:
        return text
    out = lines[:hi + 1]
    repaired = 0
    for ln in lines[hi + 1:]:
        if ln.strip() == "" or '"' in ln:
            out.append(ln)
            continue
        parts = ln.split(sep)
        surplus = len(parts) - ncol
        if surplus > 0 and idx + surplus + 1 <= len(parts):
            merged = sep.join(parts[idx: idx + surplus + 1])
            # The merged value STILL contains ``sep``, so it must be QUOTED or a
            # re-parse just re-splits it into the same surplus fields. Quote it (and
            # escape any inner quotes) so pandas' default quotechar reads it as one
            # field. Quote-bearing lines were skipped above, so this cannot clash.
            merged = '"' + merged.replace('"', '""') + '"'
            parts = parts[:idx] + [merged] + parts[idx + surplus + 1:]
            if len(parts) == ncol:
                ln = sep.join(parts)
                repaired += 1
        out.append(ln)
    if repaired:
        logger.info("tabular_parser: recovered %d over-length row(s) by merging an "
                    "unescaped %r back into column %r", repaired, sep, header[idx])
    return "\n".join(out)


def _looks_html(head: bytes) -> bool:
    s = head[:1024].lstrip().lower()
    return (s.startswith(b"<!doctype html") or s.startswith(b"<html")
            or s.startswith(b"<table") or (b"<table" in s and b"<tr" in s)
            or (s.startswith(b"<?xml") and b"spreadsheet" in s))


_MIN_STYLES = (
    b'<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
    b'<styleSheet xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main">'
    b'<fonts count="1"><font><sz val="11"/><name val="Calibri"/></font></fonts>'
    b'<fills count="1"><fill><patternFill patternType="none"/></fill></fills>'
    b'<borders count="1"><border/></borders>'
    b'<cellStyleXfs count="1"><xf numFmtId="0" fontId="0" fillId="0" borderId="0"/></cellStyleXfs>'
    b'<cellXfs count="1"><xf numFmtId="0" fontId="0" fillId="0" borderId="0" xfId="0"/></cellXfs>'
    b'<cellStyles count="1"><cellStyle name="Normal" xfId="0" builtinId="0"/></cellStyles>'
    b'</styleSheet>'
)


def _repair_xlsx(raw: bytes) -> bytes:
    """Rebuild the xlsx zip, swapping a broken xl/styles.xml for a minimal valid
    one. Fixes 'could not read stylesheet ... invalid XML' exports."""
    import io
    import zipfile
    src = zipfile.ZipFile(io.BytesIO(raw))
    buf = io.BytesIO()
    with zipfile.ZipFile(buf, "w", zipfile.ZIP_DEFLATED) as out:
        for item in src.infolist():
            data = src.read(item.filename)
            if item.filename == "xl/styles.xml":
                data = _MIN_STYLES
            out.writestr(item.filename, data)
    return buf.getvalue()


def _promote_header(raw_df: pd.DataFrame) -> pd.DataFrame:
    """Given a header-less frame, pick the most likely header row (the row with
    the most non-empty cells within the first 15) and drop the rows above it —
    handles report titles / metadata rows sitting above the real header."""
    df = raw_df.fillna("")
    n = len(df)
    if n == 0:
        return df.astype(str)
    best_row, best_filled = 0, -1
    for i in range(min(15, n)):
        filled = sum(1 for v in df.iloc[i] if str(v).strip() != "")
        if filled > best_filled:
            best_filled, best_row = filled, i
    header, seen = [], {}
    for j, v in enumerate(df.iloc[best_row]):
        name = str(v).strip() or f"col_{j + 1}"
        if name in seen:
            seen[name] += 1
            name = f"{name}_{seen[name]}"
        else:
            seen[name] = 0
        header.append(name)
    body = df.iloc[best_row + 1:].reset_index(drop=True)
    body.columns = header
    return body.astype(str)


def _read_html_table(file_path: Path) -> pd.DataFrame:
    """Some 'xlsx'/'xls'/'csv' exports (Anaplan, legacy tools) are actually HTML
    tables. Read the largest table."""
    tables = pd.read_html(str(file_path))  # requires lxml
    if not tables:
        raise ValueError("No HTML tables found in file")
    return max(tables, key=lambda d: d.shape[0] * max(1, d.shape[1])).astype(str)


def _load_xlsx_readonly(raw: bytes):
    """Open an xlsx read-only (fast, low-memory streaming), repairing a corrupt
    stylesheet on the fly if openpyxl chokes on it."""
    import io
    import openpyxl
    try:
        return openpyxl.load_workbook(io.BytesIO(raw), read_only=True, data_only=True)
    except Exception:
        return openpyxl.load_workbook(io.BytesIO(_repair_xlsx(raw)), read_only=True,
                                      data_only=True)


def _stream_sheet(ws, limit: int | None) -> list[list]:
    rows: list[list] = []
    for row in ws.iter_rows(values_only=True):
        rows.append(list(row))
        if limit is not None and len(rows) >= limit:
            break
    return rows


def _read_sheet_calamine(raw: bytes, sheet_title: str) -> pd.DataFrame:
    """Read one sheet with the calamine engine (Rust-based, ~5x faster than
    openpyxl on large workbooks). keep_default_na=False preserves literal source
    tokens like 'NULL'/'NA' as text so FBDI output stays faithful to the source."""
    import io
    return pd.read_excel(io.BytesIO(raw), sheet_name=sheet_title, engine="calamine",
                         header=None, dtype=str, keep_default_na=False)


def list_excel_sheets(file_path: Path | str) -> list[dict]:
    """List a workbook's sheets with cheap stored dimensions, largest first — used
    by the upload flow to prompt the user when a workbook has multiple data sheets
    (e.g. Customer + Address). Non-xlsx files return a single implicit sheet."""
    file_path = Path(file_path)
    raw = file_path.read_bytes()
    if raw[:4] != b"PK\x03\x04" and file_path.suffix.lower() not in (".xlsx", ".xlsm", ".xls"):
        return [{"name": file_path.stem, "rows": 0, "cols": 0}]
    try:
        wb = _load_xlsx_readonly(raw)
    except Exception:  # noqa: BLE001
        return [{"name": file_path.stem, "rows": 0, "cols": 0}]
    try:
        out = []
        for ws in wb.worksheets:
            r, c = (ws.max_row or 0), (ws.max_column or 0)
            if r == 0 and c == 0:
                continue  # skip truly empty sheets
            out.append({"name": ws.title, "rows": int(r), "cols": int(c)})
        out.sort(key=lambda s: s["rows"] * s["cols"], reverse=True)
        return out or [{"name": (wb.worksheets[0].title if wb.worksheets else file_path.stem),
                        "rows": 0, "cols": 0}]
    finally:
        try:
            wb.close()
        except Exception:  # noqa: BLE001
            pass


def _read_excel_robust(file_path: Path, raw: bytes, nrows: int | None = None,
                       sheet: str | None = None) -> pd.DataFrame:
    """Read one sheet of a workbook. By default the LARGEST sheet; pass ``sheet``
    (a title, case-insensitive) to read a specific one — the multi-sheet upload
    prompt uses this so 'Customer' and 'Address' can each become their own source.
    Full reads use the fast calamine engine; bounded samples stream with openpyxl."""
    wb = _load_xlsx_readonly(raw)
    try:
        sheets = list(wb.worksheets)
        if not sheets:
            raise ValueError("Workbook has no sheets")
        limit = (nrows + 1) if nrows else None  # +1 for the header row
        best = None
        if sheet:
            want = str(sheet).strip().lower()
            best = next((w for w in sheets if (w.title or "").strip().lower() == want), None)
        if best is None:
            # Rank sheets by stored dimensions (cheap).
            ranked = sorted(sheets, key=lambda w: (w.max_row or 0) * (w.max_column or 0),
                            reverse=True)
            best = ranked[0]
        has_dims = (best.max_row or 0) * (best.max_column or 0) > 0
        raw_df = None
        # Fast path: full read via calamine (uncapped output generation).
        if nrows is None and has_dims:
            try:
                raw_df = _read_sheet_calamine(raw, best.title)
            except Exception:  # noqa: BLE001 — engine missing / read error → stream
                raw_df = None
        if raw_df is None:  # sampled read, or calamine fallback: openpyxl streaming
            rows: list[list] = []
            if has_dims:
                rows = _stream_sheet(best, limit)
            if not rows:  # dimensions unreliable — probe each sheet, keep the biggest
                best_cells = -1
                for ws in sheets:
                    r = _stream_sheet(ws, limit)
                    cells = sum(1 for row in r for c in row if c is not None)
                    if cells > best_cells:
                        best_cells, rows = cells, r
            if not rows:
                raise ValueError("Workbook has no readable rows")
            width = max((len(r) for r in rows), default=0)
            norm = [r + [None] * (width - len(r)) for r in rows]
            raw_df = pd.DataFrame(norm)
    finally:
        try:
            wb.close()
        except Exception:  # noqa: BLE001
            pass
    return _promote_header(raw_df)


def _csv_encodings(raw: bytes) -> list[str]:
    if raw[:2] in (b"\xff\xfe", b"\xfe\xff"):
        return ["utf-16"]
    if raw[:3] == b"\xef\xbb\xbf":
        return ["utf-8-sig", "utf-8"]
    return ["utf-8", "cp1252", "latin-1"]


def _sniff_delimiter(sample: str) -> str:
    import csv as _csv
    try:
        return _csv.Sniffer().sniff(sample, delimiters=",\t;|").delimiter
    except Exception:  # noqa: BLE001
        counts = {d: sample.count(d) for d in (",", "\t", ";", "|")}
        best = max(counts, key=counts.get)
        return best if counts[best] > 0 else ","


def _read_csv_robust(file_path: Path, raw: bytes, nrows: int | None = None) -> pd.DataFrame:
    """Fast delimited-text reader: sniff encoding + delimiter, then use pandas'
    C engine (10-50x faster than the python engine on large files). Falls back to
    the python engine with delimiter auto-detection only if the C parse fails or
    collapses to a single column."""
    import io
    last_err = None
    for enc in _csv_encodings(raw):
        try:
            head = raw[:65536].decode(enc, errors="strict")
        except (UnicodeDecodeError, UnicodeError):
            continue
        first_line = head.split("\n", 1)[0]
        sep = _sniff_delimiter(first_line + "\n" + head[:8192])
        # Fast path — C engine with an explicit delimiter. Repair over-length rows
        # (an unescaped delimiter inside a free-text field) FIRST, so a real record is
        # recovered instead of dropped by on_bad_lines. Repairs on the decoded text and
        # reads from it; only the non-comma delimited exports hit this (comma CSVs are
        # normally quoted, and repairing them could fight pandas' quote handling).
        try:
            if sep != ",":
                _full = _repair_overlong_delimited(raw.decode(enc, errors="replace"), sep)
                df = pd.read_csv(io.StringIO(_full), dtype=str, keep_default_na=False,
                                 sep=sep, nrows=nrows, on_bad_lines="skip",
                                 low_memory=False)
            else:
                df = pd.read_csv(io.BytesIO(raw), dtype=str, keep_default_na=False,
                                 encoding=enc, sep=sep, nrows=nrows,
                                 on_bad_lines="skip", low_memory=False)
            if df.shape[1] > 1 or sep == ",":
                return df.astype(str)
        except (UnicodeDecodeError, UnicodeError):
            continue
        except Exception as e:  # noqa: BLE001
            last_err = e
        # Fallback — python engine sniffs the delimiter itself (slower, robust).
        try:
            df = pd.read_csv(io.BytesIO(raw), dtype=str, keep_default_na=False,
                             encoding=enc, engine="python", sep=None,
                             nrows=nrows, on_bad_lines="skip")
            return df.astype(str)
        except (UnicodeDecodeError, UnicodeError):
            continue
        except Exception as e:  # noqa: BLE001
            last_err = e
    raise ValueError(f"Could not read this file as CSV/text ({last_err or 'unknown encoding'}).")


def parse_tabular(file_path: Path | str, file_type: str | None = None,
                  nrows: int | None = None, sheet: str | None = None) -> pd.DataFrame:
    """Robust CSV / XLSX / HTML parser. Detects the real format by magic bytes
    (the extension can lie), repairs corrupt xlsx stylesheets, unwraps HTML tables
    exported as .xls(x)/.csv, sniffs CSV encoding (incl. UTF-16 / BOM) and
    delimiter, streams only the largest sheet, and detects the header row below
    titles. Pass ``nrows`` to read only the first N data rows (fast profiling)."""
    file_path = Path(file_path)
    raw = file_path.read_bytes()
    head = raw[:1024]
    if raw[:4] == b"PK\x03\x04":               # real OOXML/xlsx (zip signature)
        return _read_excel_robust(file_path, raw, nrows=nrows, sheet=sheet)
    if _looks_html(head):                        # HTML masquerading as a spreadsheet
        try:
            df = _read_html_table(file_path)
            return df.head(nrows) if nrows else df
        except Exception:  # noqa: BLE001 — fall through to text parsing
            pass
    ftype = (file_type or file_path.suffix.lstrip(".")).lower()
    if ftype in ("xlsx", "xls", "xlsm"):
        try:
            return _read_excel_robust(file_path, raw, nrows=nrows, sheet=sheet)
        except Exception:  # noqa: BLE001 — last resort: try as delimited text
            return _read_csv_robust(file_path, raw, nrows=nrows)
    return _read_csv_robust(file_path, raw, nrows=nrows)


_DATE_PATTERNS = [
    (re.compile(r"^\d{4}-\d{2}-\d{2}$"), "YYYY-MM-DD"),
    (re.compile(r"^\d{2}/\d{2}/\d{4}$"), "MM/DD/YYYY or DD/MM/YYYY"),
    (re.compile(r"^\d{4}/\d{2}/\d{2}$"), "YYYY/MM/DD"),
    (re.compile(r"^\d{2}-\d{2}-\d{4}$"), "MM-DD-YYYY or DD-MM-YYYY"),
    (re.compile(r"^\d{4}\d{2}\d{2}$"), "YYYYMMDD"),
]
_INT_RE = re.compile(r"^-?\d+$")
_FLOAT_RE = re.compile(r"^-?\d*\.\d+$")
_BOOL_VALS = {"true", "false", "yes", "no", "y", "n", "0", "1", "t", "f"}


def _classify_value(v: str) -> str:
    v = v.strip()
    if v == "":
        return "null"
    if _INT_RE.match(v):
        return "integer"
    if _FLOAT_RE.match(v):
        return "float"
    for pat, _ in _DATE_PATTERNS:
        if pat.match(v):
            return "date"
    if v.lower() in _BOOL_VALS and len(v) <= 5:
        return "boolean"
    return "string"


def _infer_column_type(values: list[str]) -> str:
    if not values:
        return "string"
    seen: dict[str, int] = {}
    for v in values:
        k = _classify_value(v)
        seen[k] = seen.get(k, 0) + 1
    seen.pop("null", None)
    if not seen:
        return "string"
    # if integers + floats → float
    if set(seen) <= {"integer", "float"}:
        return "float" if "float" in seen else "integer"
    if set(seen) == {"date"}:
        return "date"
    if set(seen) <= {"boolean"} and sum(seen.values()) > 0:
        return "boolean"
    return "string"


def _detect_pattern(values: list[str]) -> str | None:
    """Return a short human-readable pattern hint."""
    if not values:
        return None
    sample = values[: min(50, len(values))]
    for pat, label in _DATE_PATTERNS:
        if all(pat.match(v.strip()) for v in sample if v.strip()):
            return f"Date format: {label}"
    if all(_INT_RE.match(v.strip()) for v in sample if v.strip()):
        return "All numeric integers"
    if all(re.match(r"^[A-Z]{2,5}$", v.strip()) for v in sample if v.strip()):
        return "Short uppercase code (e.g. UOM, currency)"
    if all(re.match(r"^[A-Za-z0-9\-_/]+$", v.strip()) for v in sample if v.strip()):
        return "Alphanumeric identifier"
    return None


def profile_dataframe(df: pd.DataFrame) -> list[dict[str, Any]]:
    """Return a list of column-profile dicts."""
    profiles: list[dict[str, Any]] = []
    total = len(df)
    for pos, col in enumerate(df.columns):
        series = df[col].astype(str).fillna("")
        non_null = [v for v in series.tolist() if v.strip() != ""]
        nulls = total - len(non_null)
        distinct = len(set(non_null))
        sample = []
        seen_set: set[str] = set()
        for v in non_null:
            if v not in seen_set:
                seen_set.add(v)
                sample.append(v)
                if len(sample) >= 8:
                    break
        inferred = _infer_column_type(non_null)
        min_val = max_val = None
        if inferred in ("integer", "float") and non_null:
            try:
                nums = [float(v) for v in non_null if _INT_RE.match(v) or _FLOAT_RE.match(v)]
                if nums:
                    min_val = str(min(nums))
                    max_val = str(max(nums))
            except Exception:
                pass
        elif inferred == "date" and non_null:
            try:
                vals = sorted(non_null)
                min_val = vals[0]
                max_val = vals[-1]
            except Exception:
                pass
        profiles.append(
            {
                "column_name": str(col),
                "position": pos,
                "inferred_type": inferred,
                "null_count": nulls,
                "null_percent": round((nulls / total * 100) if total else 0.0, 2),
                "distinct_count": distinct,
                "sample_values": sample,
                "min_value": min_val,
                "max_value": max_val,
                "pattern_summary": _detect_pattern(non_null),
            }
        )
    return profiles
