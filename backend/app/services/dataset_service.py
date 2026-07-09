"""Dataset upload + profiling service."""
from __future__ import annotations

import re as _re
import shutil
from pathlib import Path
from typing import Any

from beanie import PydanticObjectId
from fastapi import UploadFile

from app.config import settings
from app.models.dataset import Dataset, DatasetColumnProfile
from app.models.fbdi import FBDITemplate
from app.parsers import parse_tabular, profile_dataframe


ALLOWED_DATASET_EXTS = {".csv", ".xlsx", ".xls"}

# Profile/classify on a bounded sample so uploads of multi-MB / multi-hundred-k
# row files stay fast. The FULL file is still read later for FBDI generation.
PROFILE_SAMPLE_ROWS = 3000


def count_data_rows(file_path: str | Path, file_type: str | None = None) -> int | None:
    """Count data rows (excluding the header) cheaply, without materializing the
    whole file as a profiled DataFrame. Returns None if it can't be determined."""
    import io
    p = Path(file_path)
    try:
        raw = p.read_bytes()
    except Exception:  # noqa: BLE001
        return None
    if raw[:4] == b"PK\x03\x04":  # xlsx — use stored sheet dimensions (instant)
        import openpyxl
        from app.parsers.tabular_parser import _repair_xlsx
        try:
            try:
                wb = openpyxl.load_workbook(io.BytesIO(raw), read_only=True)
            except Exception:  # noqa: BLE001
                wb = openpyxl.load_workbook(io.BytesIO(_repair_xlsx(raw)), read_only=True)
            best = max((ws.max_row or 0) for ws in wb.worksheets) if wb.worksheets else 0
            wb.close()
            return max(0, best - 1)
        except Exception:  # noqa: BLE001
            return None
    # Delimited text — count with the C parser reading a single column. This is
    # fast (<1s even at 40 MB) and, unlike counting newline bytes, is correct when
    # fields contain embedded newlines (quoted multi-line values).
    import pandas as pd
    from app.parsers.tabular_parser import _csv_encodings, _sniff_delimiter
    for enc in _csv_encodings(raw):
        try:
            head = raw[:65536].decode(enc, errors="strict")
        except (UnicodeDecodeError, UnicodeError):
            continue
        sep = _sniff_delimiter(head.split("\n", 1)[0] + "\n" + head[:8192])
        try:
            n = len(pd.read_csv(io.BytesIO(raw), dtype=str, keep_default_na=False,
                                encoding=enc, sep=sep, usecols=[0],
                                on_bad_lines="skip", low_memory=False))
            return int(n)
        except Exception:  # noqa: BLE001
            continue
    try:  # last-resort byte count
        return max(0, raw.count(b"\n") - 1)
    except Exception:  # noqa: BLE001
        return None


def save_upload(upload: UploadFile, subdir: str = "datasets") -> tuple[Path, str]:
    target_dir = settings.upload_path / subdir
    target_dir.mkdir(parents=True, exist_ok=True)
    safe_name = Path(upload.filename or "upload").name
    target = target_dir / safe_name
    counter = 1
    while target.exists():
        stem = Path(safe_name).stem
        suffix = Path(safe_name).suffix
        target = target_dir / f"{stem}_{counter}{suffix}"
        counter += 1
    with open(target, "wb") as f:
        shutil.copyfileobj(upload.file, f)
    upload.file.close()
    return target, target.name


async def create_dataset_from_upload(
    upload: UploadFile, name: str | None, description: str | None
) -> tuple[Dataset, list[DatasetColumnProfile]]:
    ext = Path(upload.filename or "").suffix.lower()
    if ext not in ALLOWED_DATASET_EXTS:
        raise ValueError(f"Unsupported file extension: {ext}")
    file_path, stored_name = save_upload(upload)

    # Dedupe: if a byte-identical file was already uploaded, reuse that dataset
    # instead of creating another copy (prevents the Datasets list filling with
    # duplicates of the same source extract).
    import hashlib
    content_hash = hashlib.sha256(Path(file_path).read_bytes()).hexdigest()
    existing = await Dataset.find_one(Dataset.content_hash == content_hash)
    if existing:
        try:
            Path(file_path).unlink(missing_ok=True)  # drop the redundant copy
        except Exception:
            pass
        cols = await DatasetColumnProfile.find(
            DatasetColumnProfile.dataset_id == existing.id
        ).sort("+position").to_list()
        return existing, cols

    # Profile on a bounded sample (fast even for 20-40 MB / 100k+ row files);
    # the true row count is read separately without materializing the whole frame.
    df = parse_tabular(file_path, file_type=ext.lstrip("."), nrows=PROFILE_SAMPLE_ROWS)
    profiles = profile_dataframe(df)
    total_rows = count_data_rows(file_path, ext.lstrip(".")) or int(len(df))

    # ── Auto-detect object type from filename + column headers ──────────────
    column_names = [p["column_name"] for p in profiles]
    templates = await FBDITemplate.find_all().to_list()
    suggestions = detect_dataset_type(stored_name, column_names, templates)
    top = suggestions[0] if suggestions else None

    ds = Dataset(
        name=name or Path(upload.filename or stored_name).stem,
        description=description,
        file_name=stored_name,
        file_path=str(file_path),
        file_type=ext.lstrip("."),
        row_count=int(total_rows),
        column_count=int(len(df.columns)),
        status="profiled",
        detected_object_type=top["business_object"] if top else None,
        detection_confidence=top["confidence"] if top else 0.0,
        detection_suggestions=suggestions,
        content_hash=content_hash,
    )
    await ds.insert()

    # Durable copy in MongoDB GridFS so output generation still works after a
    # redeploy wipes the ephemeral container disk.
    try:
        from app.services.dataset_file_store import store_dataset_bytes
        await store_dataset_bytes(ds.id, stored_name, Path(file_path).read_bytes())
    except Exception:
        pass

    # Bulk-insert column profiles in one round-trip instead of one per column
    # (200+ sequential inserts was a major upload-latency source at scale).
    col_docs = [
        DatasetColumnProfile(id=PydanticObjectId(), dataset_id=ds.id, **prof)
        for prof in profiles
    ]
    if col_docs:
        await DatasetColumnProfile.insert_many(col_docs)

    return ds, col_docs


async def get_dataset_preview(ds: Dataset, limit: int = 50) -> dict[str, Any]:
    # Rehydrate from GridFS if the ephemeral disk copy was wiped on redeploy.
    from app.services.dataset_file_store import materialize_dataset_file
    src_path = await materialize_dataset_file(ds)
    if src_path is None:
        return {"columns": [], "rows": [], "total_rows": 0, "missing_file": True}
    src_path = str(src_path)
    head = parse_tabular(src_path, file_type=ds.file_type, nrows=limit)
    total = ds.row_count or count_data_rows(src_path, ds.file_type) or int(len(head))
    return {
        "columns": list(head.columns.astype(str)),
        "rows": head.fillna("").to_dict(orient="records"),
        "total_rows": int(total),
    }


# ─── Dataset-type auto-detection ─────────────────────────────────────────────

_OBJECT_HINTS: dict[str, tuple[list[str], list[str]]] = {
    "Item": (
        ["item", "sku", "product", "material", "article", "part"],
        ["item_num", "item_number", "itemid", "uom", "u_m", "item_description",
         "item_class", "primary_uom", "product_code", "matl_type", "part_number",
         "part_type", "part_description", "product_family", "commodity",
         "revision", "lifecycle_phase", "inventory_item"],
    ),
    "Customer": (
        ["customer", "cust", "client", "buyer", "account"],
        ["customer_num", "customer_number", "cust_id", "account_num",
         "customer_name", "bill_to", "sold_to", "party_name"],
    ),
    "Supplier": (
        ["supplier", "vendor", "supp"],
        ["supplier_num", "vendor_id", "supplier_name", "vendor_name",
         "supplier_site", "payment_terms"],
    ),
    "Sales Order": (
        ["sales_order", "salesorder", "so_", "order_header"],
        ["order_number", "order_date", "customer_number", "order_type",
         "ordered_quantity", "selling_price", "ship_date", "order_status"],
    ),
    "Purchase Order": (
        ["purchase_order", "po_", "purchaseorder"],
        ["po_number", "supplier_number", "ordered_quantity",
         "need_by_date", "buyer_name", "po_line"],
    ),
    "BOM": (
        ["bom", "bill_of_material", "component"],
        ["assembly_item", "component_item", "component_quantity",
         "bom_type", "effectivity_date"],
    ),
    "On-Hand Balance": (
        ["onhand", "on_hand", "inventory", "balance", "stock"],
        ["organization_code", "item_number", "subinventory",
         "transaction_quantity", "lot_number"],
    ),
    "UOM": (
        ["uom", "unit_of_measure"],
        ["uom_code", "unit_of_measure_code", "description", "base_uom"],
    ),
    "Inventory Org": (
        ["inventory_org", "org_", "organization"],
        ["organization_code", "organization_name", "legal_entity", "location_code"],
    ),
}

_NORMALIZE = _re.compile(r"[^a-z0-9]+")


def _norm(s: str) -> str:
    return _NORMALIZE.sub("_", s.lower()).strip("_")


def _kw_score(tokens: list[str], keywords: list[str]) -> float:
    hits = sum(1 for kw in keywords if any(kw in tok for tok in tokens))
    return hits / len(keywords) if keywords else 0.0


def detect_dataset_type(filename: str, column_names: list[str], templates: list) -> list[dict]:
    fname_tokens = [_norm(t) for t in _re.split(r"[\W_]+", filename) if t]
    col_tokens = [_norm(c) for c in column_names]
    by_obj: dict[str, object] = {}
    for tpl in templates:
        obj = (tpl.business_object or tpl.name or "").strip()
        if obj and obj not in by_obj:
            by_obj[obj] = tpl
    results: list[dict] = []
    for obj, (fname_kws, col_kws) in _OBJECT_HINTS.items():
        fname_s = _kw_score(fname_tokens, fname_kws)
        col_s = _kw_score(col_tokens, col_kws)
        confidence = round(min(1.0, fname_s * 0.45 + col_s * 0.55), 3)
        if confidence < 0.10:
            continue
        tpl = by_obj.get(obj)
        if not tpl:
            continue
        reasons: list[str] = []
        if fname_s >= 0.2:
            reasons.append(f"filename contains '{obj.lower()}' keywords")
        if col_s >= 0.2:
            reasons.append(f"{int(col_s * 100)}% column-name keyword match")
        results.append({
            "template_id": str(tpl.id),
            "template_name": tpl.name,
            "business_object": obj,
            "confidence": confidence,
            "reason": "; ".join(reasons) if reasons else f"weak signal for {obj}",
        })
    results.sort(key=lambda x: -x["confidence"])
    return results[:3]


# ── Source-system detection ──────────────────────────────────────────────────
# Per-source signature keywords. We score the file name + column names against
# each source's tell-tale vocabulary. Signals are intentionally distinctive.
_SOURCE_HINTS: dict[str, tuple[list[str], list[str]]] = {
    # Oracle EBS — UPPER_SNAKE columns, *_ID surrogate keys, SEGMENT/ATTRIBUTE
    # flexfields, WHO columns, and canonical table names in the file name.
    "oracle_ebs": (
        ["mtl", "hz", "po", "oe", "ap", "ar", "gl", "ebs", "apps", "_all", "_b", "_tl"],
        ["organization_id", "inventory_item_id", "segment1", "attribute1",
         "last_update_date", "last_updated_by", "created_by", "creation_date",
         "org_id", "set_of_books_id", "party_id", "vendor_id", "person_id"],
    ),
    # NetSuite — saved-search / CSV exports lead with Internal/External ID.
    "netsuite": (
        ["netsuite", "ns", "savedsearch", "saved_search"],
        ["internal_id", "external_id", "internalid", "externalid", "nsinternal",
         "subsidiary", "netsuite_id", "name", "isinactive", "custentity", "custitem"],
    ),
    # Infor SyteLine (CloudSuite Industrial) — short lower/mixed names, site_ref,
    # cust_num / vend_num, u_m, stat.
    "syteline": (
        ["syteline", "infor", "csi", "sl_", "ebos"],
        ["item", "cust_num", "vend_num", "site_ref", "u_m", "stat",
         "description", "product_code", "unit_cost", "matl_type", "whse"],
    ),
    # Arena PLM — item-centric PLM export: Item Number, Rev, Lifecycle Phase,
    # Category, Supplier Item.
    "arena": (
        ["arena", "plm", "bom_export"],
        ["item_number", "rev", "revision", "lifecycle_phase", "category",
         "supplier_item", "manufacturer_part", "mfr_part", "guid", "effectivity"],
    ),
}

_SOURCE_DISPLAY = {
    "oracle_ebs": "Oracle EBS", "netsuite": "NetSuite",
    "syteline": "Infor SyteLine", "arena": "Arena PLM", "custom": "Custom / Other",
}


def detect_source_system(filename: str, column_names: list[str]) -> list[dict]:
    """Rank likely source systems for an uploaded file by filename + columns."""
    fname_tokens = [_norm(t) for t in _re.split(r"[\W_]+", filename) if t]
    col_tokens = [_norm(c) for c in column_names]
    out: list[dict] = []
    for code, (fname_kws, col_kws) in _SOURCE_HINTS.items():
        fs = _kw_score(fname_tokens, fname_kws)
        cs = _kw_score(col_tokens, col_kws)
        conf = round(min(1.0, fs * 0.55 + cs * 0.75), 3)
        reasons = []
        if fs >= 0.15:
            reasons.append("file name matches its naming convention")
        if cs >= 0.15:
            reasons.append(f"{int(cs * 100)}% signature-column match")
        out.append({
            "code": code,
            "display": _SOURCE_DISPLAY.get(code, code),
            "confidence": conf,
            "reason": "; ".join(reasons) if reasons else "weak signal",
        })
    out.sort(key=lambda x: -x["confidence"])
    return out


def column_signature(column_names: list[str]) -> str:
    """Stable signature of a file's column set — the learning key. Order- and
    case-insensitive, so re-exports of the 'same' file hit the learned record."""
    import hashlib
    norm = sorted({_norm(c) for c in column_names if c})
    return hashlib.sha1("|".join(norm).encode("utf-8")).hexdigest()[:16]
