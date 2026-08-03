"""The run report: what the tool actually did to the input file.

An analyst downloads a bundle and gets a set of FBDI files. This answers the
question that immediately follows — *what happened between the raw extract and
those files?* Which columns were mapped and on whose authority, what was
cleansed, what was merged away as duplicate, what validation found, and which
required fields are still short.

Two rules shaped it.

**It reports what the run DID, not what a fresh recompute says now.** Every
number here comes from the artifact the bundle actually contains — the
``dq_report`` persisted on ``ConvertedOutput``, the mapping rows as they stood,
the row counts of the files on disk. Recomputing cleansing or validation at
report time would produce a document that quietly disagrees with the files it
describes, which is the exact failure this codebase keeps paying for. Where a
figure cannot be sourced from the run it is left blank and says so, rather than
being filled in from somewhere else.

**Every section names its authority.** A mapping is not just "Supplier Name ←
vendor_name"; it is that, decided by the mapping workbook on 31-Jul, and now a
view of the store rather than a decision in its own right. A report that says
what happened without saying who said so cannot be argued with, and the
arguments are the point.

Pure: this module takes dicts and returns bytes. No database, no I/O beyond the
in-memory workbook, so it is unit-testable against hand-built input.
"""
from __future__ import annotations

import io
from datetime import datetime
from typing import Any, Iterable, Optional

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

# ── House style ──────────────────────────────────────────────────────────────
# Deliberately the same palette and weights as mapping_export_service, so a
# client receiving both does not get two different-looking documents.

_INK = "1F2937"
_MUTED = "6B7280"
_BRAND = "1E3A5F"
_RULE = "D1D5DB"

_OK = "047857"
_WARN = "B45309"
_BAD = "B91C1C"

_FILL_HEAD = PatternFill("solid", fgColor=_BRAND)
_FILL_BAND = PatternFill("solid", fgColor="F3F4F6")
_FILL_OK = PatternFill("solid", fgColor="ECFDF5")
_FILL_WARN = PatternFill("solid", fgColor="FFFBEB")
_FILL_BAD = PatternFill("solid", fgColor="FEF2F2")

_F_TITLE = Font(name="Calibri", size=16, bold=True, color=_BRAND)
_F_SUB = Font(name="Calibri", size=10, color=_MUTED)
_F_HEAD = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
_F_BODY = Font(name="Calibri", size=10, color=_INK)
_F_MUTED = Font(name="Calibri", size=10, color=_MUTED)
_F_STRONG = Font(name="Calibri", size=10, bold=True, color=_INK)
_F_NOTE = Font(name="Calibri", size=9, color=_MUTED, italic=True)

_THIN = Side(style="thin", color=_RULE)
_BORDER = Border(bottom=_THIN)
_WRAP = Alignment(vertical="top", wrap_text=True)
_TOP = Alignment(vertical="top")


def _sheet(wb: Workbook, name: str, title: str, blurb: str) -> Any:
    ws = wb.create_sheet(name[:31])
    ws.sheet_view.showGridLines = False
    ws["A1"] = title
    ws["A1"].font = _F_TITLE
    ws["A2"] = blurb
    ws["A2"].font = _F_SUB
    ws.row_dimensions[1].height = 22
    return ws


def _headers(ws, row: int, headers: list[tuple[str, int]]) -> None:
    for i, (label, width) in enumerate(headers, start=1):
        c = ws.cell(row=row, column=i, value=label)
        c.font = _F_HEAD
        c.fill = _FILL_HEAD
        c.alignment = Alignment(vertical="center", wrap_text=True)
        ws.column_dimensions[get_column_letter(i)].width = width
    ws.row_dimensions[row].height = 24
    ws.freeze_panes = ws.cell(row=row + 1, column=1)


def _row(ws, row: int, values: Iterable[Any], *, band: bool = False,
         fill: Optional[PatternFill] = None, strong_first: bool = False) -> None:
    for i, v in enumerate(values, start=1):
        c = ws.cell(row=row, column=i, value=v)
        c.font = _F_STRONG if (strong_first and i == 1) else _F_BODY
        c.alignment = _WRAP
        c.border = _BORDER
        if fill is not None:
            c.fill = fill
        elif band:
            c.fill = _FILL_BAND


def _empty(ws, row: int, message: str) -> None:
    c = ws.cell(row=row, column=1, value=message)
    c.font = _F_NOTE


def _autofilter(ws, first_row: int, last_row: int, cols: int) -> None:
    if last_row > first_row:
        ws.auto_filter.ref = f"A{first_row}:{get_column_letter(cols)}{last_row}"


def _when(value: Any) -> str:
    if isinstance(value, datetime):
        return value.strftime("%d-%b-%Y %H:%M")
    return str(value or "")


def _n(value: Any, blank: str = "—") -> Any:
    """A number, or a dash. Never a zero standing in for "we do not know"."""
    return blank if value is None else value


# ── Sections ─────────────────────────────────────────────────────────────────

def _summary(wb: Workbook, title: str, objects: list[dict], generated_at: Any) -> None:
    ws = wb.create_sheet("Summary")
    ws.sheet_view.showGridLines = False
    ws["A1"] = title
    ws["A1"].font = _F_TITLE
    ws["A2"] = (f"What the tool did to the input files. Generated {_when(generated_at)}. "
                "Every figure comes from the run that produced the bundle, not from a "
                "fresh recalculation — so this document and those files agree.")
    ws["A2"].font = _F_SUB
    ws.row_dimensions[1].height = 22
    ws.row_dimensions[2].height = 28
    ws.merge_cells("A2:J2")

    headers = [("Interface object", 26), ("Source rows", 12), ("Output rows", 12),
               ("Merged / de-duplicated", 20), ("Columns mapped", 15),
               ("Cleansed values", 15), ("Validation errors", 15),
               ("Warnings", 11), ("Required short", 14), ("Output file", 34)]
    _headers(ws, 4, headers)

    r = 5
    for i, o in enumerate(objects):
        blocked = bool(o.get("blocked"))
        short = int(o.get("required_failed") or 0)
        fill = _FILL_BAD if (blocked or short) else (
            _FILL_WARN if int(o.get("warning_count") or 0) else None)
        _row(ws, r, [
            o.get("object") or "(unnamed)",
            _n(o.get("source_rows")),
            _n(o.get("output_rows")),
            _n(o.get("merged_or_deduped")),
            f'{o.get("mapped", 0)} of {o.get("total_fields", 0)}',
            _n(o.get("cleansing_fix_count")),
            int(o.get("error_count") or 0),
            int(o.get("warning_count") or 0),
            short,
            o.get("output_file") or "not generated",
        ], band=(i % 2 == 1), fill=fill, strong_first=True)
        r += 1
    if not objects:
        _empty(ws, r, "No conversions in this project have been generated yet.")
        r += 1

    r += 1
    ws.cell(row=r, column=1, value="How to read this").font = _F_STRONG
    r += 1
    for line in (
        "Source rows — records read from the uploaded extract(s), before anything was done to them.",
        "Merged / de-duplicated — records that did not reach the output because a natural key "
        "already had a surviving row. Blank means the run cannot account for the difference.",
        "Columns mapped — target fields that resolved to a source column, a constant or a rule. "
        "The Mappings sheet says which, and on whose authority.",
        "Cleansed values — individual values the tool changed. The Cleansing sheet shows before "
        "and after for each rule that fired.",
        "Validation errors — advisory. A red row is one Oracle is likely to reject on load.",
        "Required short — curated required fields that are blank or partly blank in the output.",
    ):
        c = ws.cell(row=r, column=1, value=line)
        c.font = _F_MUTED
        c.alignment = _WRAP
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=10)
        ws.row_dimensions[r].height = 26
        r += 1


def _mappings(wb: Workbook, rows: list[dict]) -> None:
    ws = _sheet(wb, "Mappings", "Column mappings",
                "Every target field, what fed it, and who decided. "
                "\"Decided by\" is the layer that resolved the field; \"authority\" is the "
                "dated entry behind it, where the tool recorded one.")
    headers = [("Object", 22), ("Interface sheet", 22), ("Target field", 30),
               ("Source column", 28), ("Constant", 18), ("Transformation", 20),
               ("Decided by", 20), ("Authority", 38), ("Status", 14),
               ("Confidence", 11), ("Required", 10)]
    _headers(ws, 4, headers)
    r = 5
    for i, m in enumerate(rows):
        layer = m.get("layer_label") or ""
        fill = None
        if m.get("required") and m.get("layer") == "unmapped":
            fill = _FILL_BAD
        elif m.get("layer") in ("ai", "deterministic"):
            fill = _FILL_WARN
        _row(ws, r, [
            m.get("object"), m.get("sheet"), m.get("target_field"),
            m.get("source_column"), m.get("default_value"),
            m.get("rule_type"), layer, m.get("authority"),
            m.get("status"),
            (round(float(m["confidence"]), 2) if m.get("confidence") is not None else None),
            "Yes" if m.get("required") else "",
        ], band=(i % 2 == 1), fill=fill)
        r += 1
    if not rows:
        _empty(ws, r, "No mappings recorded.")
        r += 1
    _autofilter(ws, 4, r - 1, len(headers))


def _cleansing(wb: Workbook, rows: list[dict]) -> None:
    ws = _sheet(wb, "Cleansing", "Cleansing performed",
                "Values the tool changed on the way through, one line per rule per field, "
                "with real examples taken from this run. Case and legal-suffix "
                "standardisation are off unless someone switched them on, because they "
                "rewrite business values.")
    headers = [("Object", 22), ("Field", 30), ("Rule", 34), ("Values changed", 14),
               ("Example before", 34), ("Example after", 34)]
    _headers(ws, 4, headers)
    r = 5
    for i, c in enumerate(rows):
        _row(ws, r, [c.get("object"), c.get("field"),
                     c.get("label") or c.get("rule"), c.get("count"),
                     c.get("before"), c.get("after")], band=(i % 2 == 1))
        r += 1
    if not rows:
        _empty(ws, r, "Nothing needed cleansing in this run.")
        r += 1
    _autofilter(ws, 4, r - 1, len(headers))


def _duplicates(wb: Workbook, rows: list[dict]) -> None:
    ws = _sheet(wb, "Duplicates", "Duplicates and merging",
                "What happened when several source records resolved to the same entity. "
                "Counts are the difference between what went in and what came out of this "
                "run, so they describe the delivered files rather than a later re-scan.")
    headers = [("Object", 24), ("Source files", 34), ("Source rows", 13),
               ("Output rows", 13), ("Merged / de-duplicated", 20),
               ("Natural key", 30), ("Analyst decisions recorded", 22)]
    _headers(ws, 4, headers)
    r = 5
    for i, d in enumerate(rows):
        removed = d.get("merged_or_deduped")
        fill = _FILL_OK if (removed or 0) else None
        _row(ws, r, [d.get("object"), d.get("sources"), _n(d.get("source_rows")),
                     _n(d.get("output_rows")), _n(removed), d.get("key"),
                     _n(d.get("decisions"), 0)], band=(i % 2 == 1), fill=fill)
        r += 1
    if not rows:
        _empty(ws, r, "No generated output to reconcile.")
        r += 1
    _autofilter(ws, 4, r - 1, len(headers))


def _validation(wb: Workbook, rows: list[dict]) -> None:
    ws = _sheet(wb, "Validation", "Validation findings",
                "What the tool checked the output against, and what it found. Advisory: a "
                "finding does not stop the download, it tells you what Oracle is likely to "
                "say. The first 50 findings per object are recorded on the artifact.")
    headers = [("Object", 22), ("Severity", 11), ("Issue", 30), ("Field", 28),
               ("Rows affected", 13), ("What it means", 46), ("Suggested fix", 40)]
    _headers(ws, 4, headers)
    r = 5
    for i, v in enumerate(rows):
        sev = str(v.get("severity") or "").lower()
        fill = _FILL_BAD if sev == "error" else (_FILL_WARN if sev == "warning" else None)
        _row(ws, r, [v.get("object"), (v.get("severity") or "").title(),
                     v.get("issue_type"), v.get("field_name"),
                     v.get("impacted_count"), v.get("message"),
                     v.get("suggested_fix")], band=(i % 2 == 1), fill=fill)
        r += 1
    if not rows:
        _empty(ws, r, "No validation findings — every checked value passed.")
        r += 1
    _autofilter(ws, 4, r - 1, len(headers))


def _required(wb: Workbook, rows: list[dict]) -> None:
    ws = _sheet(wb, "Required fields", "Required fields",
                "Oracle's mandatory columns for the interfaces this bundle writes, checked "
                "against the generated sheets. \"Not owned\" means the curated list covers a "
                "sheet this conversion does not produce, so nothing was checked there.")
    headers = [("Object", 22), ("Interface sheet", 30), ("Field", 32),
               ("Result", 14), ("Detail", 46)]
    _headers(ws, 4, headers)
    r = 5
    for i, q in enumerate(rows):
        status = str(q.get("status") or "").lower()
        fill = (_FILL_BAD if status == "failed"
                else _FILL_WARN if status == "partial" else None)
        _row(ws, r, [q.get("object"), q.get("sheet"), q.get("field"),
                     (q.get("status") or "").title(), q.get("detail")],
             band=(i % 2 == 1), fill=fill)
        r += 1
    if not rows:
        _empty(ws, r, "Every curated required field is populated.")
        r += 1
    _autofilter(ws, 4, r - 1, len(headers))


def _run_log(wb: Workbook, rows: list[dict]) -> None:
    ws = _sheet(wb, "Run log", "Run log",
                "One line per generated artifact: when it was built, from what, and "
                "anything that went wrong while building it. A failure recorded here means "
                "the file shipped anyway, without that step — it is not hidden.")
    headers = [("Object", 22), ("Output file", 38), ("Generated", 20),
               ("Rows", 10), ("Columns", 10), ("Store decisions applied", 18),
               ("State", 14), ("Problem during generation", 48)]
    _headers(ws, 4, headers)
    r = 5
    for i, g in enumerate(rows):
        problem = g.get("problem")
        stale = str(g.get("state") or "").lower() == "stale"
        fill = _FILL_BAD if problem else (_FILL_WARN if stale else None)
        _row(ws, r, [g.get("object"), g.get("output_file"), _when(g.get("generated_at")),
                     _n(g.get("rows")), _n(g.get("columns")),
                     _n(g.get("learnings_applied")),
                     (g.get("state") or "").title(), problem or ""],
             band=(i % 2 == 1), fill=fill)
        r += 1
    if not rows:
        _empty(ws, r, "Nothing has been generated for this project yet.")
        r += 1
    _autofilter(ws, 4, r - 1, len(headers))


# ── The one entry point ──────────────────────────────────────────────────────

SECTIONS = ("summary", "mappings", "cleansing", "duplicates", "validation",
            "required", "run_log")


def build_workbook(*, title: str, generated_at: Any = None,
                   objects: Optional[list[dict]] = None,
                   mappings: Optional[list[dict]] = None,
                   cleansing: Optional[list[dict]] = None,
                   duplicates: Optional[list[dict]] = None,
                   validation: Optional[list[dict]] = None,
                   required: Optional[list[dict]] = None,
                   run_log: Optional[list[dict]] = None) -> bytes:
    """The run report as an .xlsx, returned as bytes.

    Every argument is a plain list of dicts the caller has already gathered, so
    this can be tested against hand-built input and so the collector can be read
    on its own. A section with nothing in it still gets a sheet, carrying a line
    saying so — an absent sheet and a section that found nothing look identical
    otherwise, and that ambiguity is how "the tool did not run it" gets read as
    "there was nothing to find".
    """
    wb = Workbook()
    wb.remove(wb.active)
    _summary(wb, title, objects or [], generated_at or datetime.utcnow())
    _mappings(wb, mappings or [])
    _cleansing(wb, cleansing or [])
    _duplicates(wb, duplicates or [])
    _validation(wb, validation or [])
    _required(wb, required or [])
    _run_log(wb, run_log or [])
    bio = io.BytesIO()
    wb.save(bio)
    return bio.getvalue()
