"""Populate the REAL Oracle FBDI Excel template with converted data.

The tool ships the actual Oracle FBDI/CSV-generation workbooks (``app/data/
fbdi_templates/*.xlsm``). This service opens one of those workbooks, finds each
interface sheet's header + first data row, wipes the shipped SAMPLE rows, and
writes the converted rows into the correct columns — so the deliverable is the
Oracle-provided template *filled in* (macros/instructions/formatting intact),
not a fresh workbook. Analysts can open it, review, and run the template's own
"Generate CSV" macro to produce the load files.

Two template layouts are handled, matching ``parsers/fbdi_parser.py`` exactly so
the fill columns line up with the parsed :class:`FBDIField` columns:

* Standard tabular (Supplier / BOM / Customer): a title row, a "* Required"
  legend, then the header row (most-populated of the first 15 rows), data from
  the next row, fields from column A.
* Oracle transposed (Item): column A carries row labels
  (Name / Description / Data Type / Technical Name …); the field headers are the
  "Name" row from column B onward, and data starts below the last label row.

Kept dependency-light (openpyxl + pandas) so the column placement can be unit
tested without the Beanie/Mongo stack.
"""
from __future__ import annotations

import io
import re
from pathlib import Path
from typing import Optional

import pandas as pd
from openpyxl import load_workbook


def _norm(s) -> str:
    """Normalise a header for matching: drop '*' required markers, whitespace and
    punctuation, lowercase. 'Import Action *' and 'Import Action' both collapse to
    'importaction'."""
    return re.sub(r"[^a-z0-9]", "", str(s if s is not None else "").lower())


def _nonempty_count(row) -> int:
    return sum(1 for v in row if v is not None and str(v).strip() != "")


def detect_layout(ws, scan: int = 15) -> tuple[int, int, int]:
    """Return ``(header_row, data_start_row, col_start)`` (all 1-based) for a data
    sheet, using the same detection as the FBDI parser so fills align with the
    stored field columns."""
    rows: list[tuple] = []
    for i, row in enumerate(ws.iter_rows(min_row=1, max_row=scan, values_only=True)):
        rows.append(row)
    # Column-A labels in the scan window.
    col_a = [(i + 1, str(rows[i][0]).strip())
             for i in range(len(rows))
             if rows[i] and rows[i][0] is not None and str(rows[i][0]).strip() != ""]
    is_oracle = any(lbl.lower() == "name" for _, lbl in col_a)
    if is_oracle:
        name_row = next((r for r, lbl in col_a if lbl.lower() == "name"), 1)
        # Data begins after the LAST metadata label row (Name/Description/Data
        # Type/Reserved/Technical Name …), which are the only non-empty col-A rows.
        last_label = max((r for r, _ in col_a), default=name_row)
        return name_row, last_label + 1, 2
    # Tabular: header = the most-populated of the first `scan` rows (earliest on tie).
    header_idx, best = 0, -1
    for i, row in enumerate(rows):
        cnt = _nonempty_count(row)
        if cnt > best:
            best, header_idx = cnt, i
    header_row = header_idx + 1
    return header_row, header_row + 1, 1


def _clean_cell(val):
    """Excel cell value for a converted DataFrame cell — blanks become empty."""
    if val is None:
        return None
    if isinstance(val, float) and pd.isna(val):
        return None
    s = val if isinstance(val, str) else str(val)
    return None if s.strip() == "" else val


def fill_template(src_path: str | Path, frames_by_sheet: dict[str, "pd.DataFrame"]) -> bytes:
    """Open the Oracle template at ``src_path`` and write each sheet's frame into
    the matching worksheet, clearing the shipped sample rows first. Sheets absent
    from ``frames_by_sheet`` (e.g. the Instructions sheet) are left untouched.
    Returns the populated workbook bytes (.xlsm preserved with macros)."""
    src_path = Path(src_path)
    keep_vba = src_path.suffix.lower() == ".xlsm"
    wb = load_workbook(src_path, keep_vba=keep_vba)
    try:
        by_norm_sheet = { _norm(name): name for name in frames_by_sheet }
        for ws_name in wb.sheetnames:
            if ws_name.lower().startswith("instruction"):
                continue
            key = _norm(ws_name)
            src_name = by_norm_sheet.get(key)
            if src_name is None:
                continue
            df = frames_by_sheet[src_name]
            ws = wb[ws_name]
            header_row, data_start, col_start = detect_layout(ws)
            max_col = ws.max_column or 0
            # header text (normalised) -> template column index
            hdr_to_col: dict[str, int] = {}
            for c in range(col_start, max_col + 1):
                v = ws.cell(row=header_row, column=c).value
                if v is None or str(v).strip() == "":
                    continue
                hdr_to_col.setdefault(_norm(v), c)
            if not hdr_to_col:
                continue
            # Which df column feeds each template column (by normalised header).
            placements: list[tuple[int, int]] = []  # (df_col_pos, template_col_idx)
            if df is not None and len(df.columns):
                for pos, dfcol in enumerate(df.columns):
                    ci = hdr_to_col.get(_norm(dfcol))
                    if ci:
                        placements.append((pos, ci))
            # Wipe shipped SAMPLE rows in every mapped column so no demo data leaks
            # through (and empty child tabs end up truly empty).
            last_row = ws.max_row or data_start
            clear_cols = set(hdr_to_col.values())
            for r in range(data_start, last_row + 1):
                for c in clear_cols:
                    ws.cell(row=r, column=c).value = None
            # Write the converted rows.
            if placements and df is not None and len(df):
                values = df.values
                for i in range(len(values)):
                    rr = data_start + i
                    row = values[i]
                    for pos, ci in placements:
                        cv = _clean_cell(row[pos])
                        if cv is not None:
                            ws.cell(row=rr, column=ci, value=cv)
        bio = io.BytesIO()
        wb.save(bio)
        return bio.getvalue()
    finally:
        wb.close()
