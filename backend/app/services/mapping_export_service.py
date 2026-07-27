"""Banded field-mapping export — the analyst-friendly workbook (Issue #3).

Turns a conversion's per-field mapping suggestions into a clean Excel workbook:
a Summary sheet (field count + description per confidence band) plus one sheet per
NON-EMPTY band, each listing Target FBDI Field / Suggested Source Field / Confidence %
/ Reason / Excluded (Do-Not-Map Rule). Matches the reference
``Item_NetSuite_Field_Mapping_Clean.xlsx``.

``build_workbook`` is a self-contained openpyxl builder (unit-testable); the router
passes it the rows the Mapping Review screen already computed, so the export always
matches what the analyst sees.
"""
from __future__ import annotations

import io
import re

# (key, label, lo, hi, sheet_name, description) — hi is exclusive except the 100 band.
BANDS = [
    ("exact",  "100% - Exact Match", 100, 101, "100pct_-_Exact_Match",   "Learned / auto-applied mapping"),
    ("b95",    "95-100%",             95, 100, "95-100pct",               "Very high confidence suggestion"),
    ("b90",    "90-95%",              90,  95, "90-95pct",                "High confidence suggestion"),
    ("b85",    "85-90%",              85,  90, "85-90pct",                "Good confidence suggestion"),
    ("b75",    "75-85%",              75,  85, "75-85pct",                "Medium confidence suggestion — review recommended"),
    ("b50",    "50-75%",              50,  75, "50-75pct",                "Low confidence suggestion — review required"),
    ("b0",     "0-50%",                1,  50, "0-50pct",                 "Very low confidence suggestion — likely needs manual mapping"),
    ("none",   "0% - No Match Found",  0,   1, "0pct_-_No_Match_Found",   "No plausible source field identified"),
]
_HEADERS = ["Target FBDI Field", "Suggested Source Field", "Confidence %", "Reason",
            "Excluded (Do-Not-Map Rule)", "Vetted Alternatives (AI-checked)",
            "Value Crosswalks (legacy → Oracle)"]

# Flat "All Fields" sheet — one row per field, original order, mirroring the columns
# analysts were used to in the old CSV export PLUS the AI-vetted reason columns.
_FLAT_HEADERS = ["Source Field", "Target FBDI Field", "Required", "How it's mapped",
                 "Transform", "Confidence %", "Status", "Needs confirmation", "Why",
                 "Other options (AI-vetted, with reasons)",
                 "Value Crosswalks (legacy → Oracle)", "Excluded (Do-Not-Map Rule)",
                 "Notes"]


def _fmt_alternatives(alts) -> str:
    """One line per ranked source candidate: ``source (72%) — accept: reason``.
    ``alts``: [{source, confidence(0-100|0..1|None), verdict?, reason?}]."""
    if not isinstance(alts, list):
        return ""
    lines = []
    for a in alts:
        if not isinstance(a, dict):
            continue
        src = str(a.get("source") or a.get("source_column") or "").strip()
        if not src:
            continue
        c = a.get("confidence")
        try:
            cf = float(c)
            if cf <= 1.0:
                cf *= 100.0
            pct = f" ({int(round(cf))}%)"
        except (TypeError, ValueError):
            pct = ""
        verdict = str(a.get("verdict") or a.get("ai_verdict") or "").strip()
        reason = str(a.get("reason") or a.get("ai_reason") or "").strip()
        tail = ""
        if verdict:
            tail = f" — {verdict}" + (f": {reason}" if reason else "")
        elif reason:
            tail = f" — {reason}"
        lines.append(f"{src}{pct}{tail}")
    return "\n".join(lines)


def _fmt_crosswalks(cws) -> str:
    """One line per value pair: ``legacy → ORACLE_CODE (vetted)``.
    ``cws``: [{legacy, oracle, status?}]."""
    if not isinstance(cws, list):
        return ""
    lines = []
    for c in cws:
        if not isinstance(c, dict):
            continue
        legacy = str(c.get("legacy") or c.get("from") or c.get("source_value") or "").strip()
        oracle = str(c.get("oracle") or c.get("to") or c.get("target_value") or "").strip()
        if not legacy and not oracle:
            continue
        status = str(c.get("status") or "").strip()
        lines.append(f"{legacy} → {oracle}" + (f"  ({status})" if status else ""))
    return "\n".join(lines)


def band_for(conf) -> str:
    """Band key for a confidence percentage (0-100). None/blank → no-match."""
    try:
        c = int(round(float(conf)))
    except (TypeError, ValueError):
        c = 0
    if c >= 100:
        return "exact"
    if c <= 0:
        return "none"
    for key, _lbl, lo, hi, _sn, _d in BANDS:
        if key in ("exact", "none"):
            continue
        if lo <= c < hi:
            return key
    return "none"


def _style():
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    thin = Side(style="thin", color="C7D2E0")
    return {
        "title": Font(name="Calibri", size=14, bold=True, color="1F4E79"),
        "hdr": Font(name="Calibri", size=11, bold=True, color="FFFFFF"),
        "hdr_fill": PatternFill("solid", fgColor="2E75B6"),
        "cell": Font(name="Calibri", size=10),
        "bold": Font(name="Calibri", size=11, bold=True),
        "border": Border(left=thin, right=thin, top=thin, bottom=thin),
        "wrap": Alignment(vertical="top", wrap_text=True),
        "excl": Font(name="Calibri", size=10, bold=True, color="C0392B"),
    }


def build_workbook(title: str, records: list[dict]) -> bytes:
    """``records``: [{target_field, suggested_source, confidence(0-100|None),
    reason, excluded(bool), alternatives, crosswalks, required, how_mapped,
    transform, status, needs_confirmation, notes}]. The extra fields feed the flat
    "All Fields" sheet and are all optional. Returns the .xlsx bytes."""
    from openpyxl import Workbook
    from openpyxl.styles import PatternFill
    st = _style()

    # Bucket records by band, preserving input order.
    buckets: dict[str, list[dict]] = {k: [] for k, *_ in BANDS}
    for r in records:
        buckets[band_for(r.get("confidence"))].append(r)

    wb = Workbook()

    # ---- Summary ----
    s = wb.active
    s.title = "Summary"
    s.sheet_view.showGridLines = False
    s["A1"] = title
    s["A1"].font = st["title"]
    s.merge_cells("A1:C1")
    for i, h in enumerate(["Confidence Band", "Field Count", "Description"]):
        c = s.cell(row=3, column=1 + i, value=h)
        c.font = st["hdr"]; c.fill = st["hdr_fill"]; c.border = st["border"]
    row = 4
    for key, label, _lo, _hi, _sn, desc in BANDS:
        s.cell(row=row, column=1, value=label).font = st["cell"]
        s.cell(row=row, column=2, value=len(buckets[key])).font = st["cell"]
        s.cell(row=row, column=3, value=desc).font = st["cell"]
        for col in range(1, 4):
            s.cell(row=row, column=col).border = st["border"]
        row += 1
    s.cell(row=row, column=1, value="Total").font = st["bold"]
    s.cell(row=row, column=2, value=len(records)).font = st["bold"]
    for col in range(1, 4):
        s.cell(row=row, column=col).border = st["border"]
    s.column_dimensions["A"].width = 24
    s.column_dimensions["B"].width = 14
    s.column_dimensions["C"].width = 52

    # ---- All Fields (flat) — the familiar single-table view, one row per field ----
    fs = wb.create_sheet(title="All Fields")
    fs.sheet_view.showGridLines = False
    fs["A1"] = f"{title}  —  all {len(records)} field{'s' if len(records) != 1 else ''}"
    fs["A1"].font = st["title"]
    fs.merge_cells("A1:M1")
    for i, h in enumerate(_FLAT_HEADERS):
        c = fs.cell(row=2, column=1 + i, value=h)
        c.font = st["hdr"]; c.fill = st["hdr_fill"]; c.border = st["border"]
        c.alignment = st["wrap"]
    fr = 3
    for rec in records:
        conf = rec.get("confidence")
        vals = [
            rec.get("suggested_source") or "",
            rec.get("target_field") or "",
            "Yes" if rec.get("required") else "",
            rec.get("how_mapped") or "",
            rec.get("transform") or "",
            ("" if conf is None else int(round(float(conf)))),
            rec.get("status") or "",
            rec.get("needs_confirmation") or "",
            rec.get("reason") or "",
            _fmt_alternatives(rec.get("alternatives")),
            _fmt_crosswalks(rec.get("crosswalks")),
            "Yes" if rec.get("excluded") else "",
            rec.get("notes") or "",
        ]
        for i, v in enumerate(vals):
            cell = fs.cell(row=fr, column=1 + i, value=v)
            cell.border = st["border"]; cell.alignment = st["wrap"]; cell.font = st["cell"]
        if rec.get("excluded"):
            fs.cell(row=fr, column=12).font = st["excl"]
        fr += 1
    for col, w in zip("ABCDEFGHIJKLM",
                      (26, 30, 10, 20, 16, 11, 14, 18, 42, 50, 40, 12, 40)):
        fs.column_dimensions[col].width = w
    fs.freeze_panes = "A3"

    # ---- One sheet per non-empty band ----
    for key, label, _lo, _hi, sheet_name, _desc in BANDS:
        rows = buckets[key]
        if not rows:
            continue
        ws = wb.create_sheet(title=sheet_name[:31])
        ws.sheet_view.showGridLines = False
        ws["A1"] = f"Confidence Band: {label}  ({len(rows)} field{'s' if len(rows) != 1 else ''})"
        ws["A1"].font = st["title"]
        ws.merge_cells("A1:G1")
        for i, h in enumerate(_HEADERS):
            c = ws.cell(row=2, column=1 + i, value=h)
            c.font = st["hdr"]; c.fill = st["hdr_fill"]; c.border = st["border"]
        r = 3
        for rec in rows:
            ws.cell(row=r, column=1, value=rec.get("target_field") or "")
            ws.cell(row=r, column=2, value=rec.get("suggested_source") or "")
            conf = rec.get("confidence")
            ws.cell(row=r, column=3, value=("" if conf is None else int(round(float(conf)))))
            ws.cell(row=r, column=4, value=rec.get("reason") or "")
            excl = ws.cell(row=r, column=5, value="Yes" if rec.get("excluded") else "")
            ws.cell(row=r, column=6, value=_fmt_alternatives(rec.get("alternatives")))
            ws.cell(row=r, column=7, value=_fmt_crosswalks(rec.get("crosswalks")))
            for col in range(1, 8):
                cell = ws.cell(row=r, column=col)
                cell.border = st["border"]
                cell.alignment = st["wrap"]
                cell.font = st["cell"]
            if rec.get("excluded"):
                excl.font = st["excl"]
            r += 1
        for col, w in zip("ABCDEFG", (30, 28, 11, 46, 14, 46, 40)):
            ws.column_dimensions[col].width = w
        ws.freeze_panes = "A3"

    bio = io.BytesIO()
    wb.save(bio)
    return bio.getvalue()
