"""Every duplicate suspect group as a spreadsheet — not just the ones on screen.

WHY THIS EXISTS
---------------
The panel returns at most ``max_clusters`` groups. On the NextPower supplier run
that is 100 of 342, and the banner said so honestly — "242 more are not listed
and cannot be decided here … raise the limit to review them all" — while offering
no control that raises it. A message telling somebody to do a thing the screen
cannot do is worse than no message: it reads as their oversight.

Two answers, and this is the second one. The limit is now settable on the panel,
which makes the other 242 decidable. But 342 groups over 831 records is not a
thing anyone reviews by scrolling, and the question actually being asked of it —
"which of these groups disagree on a strong identifier, so which are probably
NOT one entity" — is a filter and a sort, which is what a spreadsheet is for.

WHAT IT REPORTS, AND WHAT IT DOES NOT
-------------------------------------
One row per member record, grouped, with the group's match confidence, the
columns the match was made on, any recorded verdict, and — the column this was
built for — the strong identifiers the group DISAGREES on.

``id_conflicts`` is advisory here exactly as it is in the panel. A differing tax
id usually means separate legal entities, but it can equally be one source
recording the number wrong, and auto-splitting the cluster would hide a real
duplicate. So the disagreement is reported and the analyst decides. That
judgement does not move into a spreadsheet just because the spreadsheet is
easier to read.

Pure: dicts in, bytes out. Same palette and shape as
``conversion_report_service``, deliberately, so the two files look like they came
from the same tool.
"""
from __future__ import annotations

from io import BytesIO
from typing import Any, Iterable

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

_BRAND = "4F46E5"
_INK = "111827"
_MUTED = "6B7280"

_F_TITLE = Font(name="Calibri", size=15, bold=True, color=_INK)
_F_SUB = Font(name="Calibri", size=10, color=_MUTED)
_F_HEAD = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
_F_BODY = Font(name="Calibri", size=10, color=_INK)
_F_GROUP = Font(name="Calibri", size=10, bold=True, color=_INK)

_FILL_HEAD = PatternFill("solid", fgColor=_BRAND)
_FILL_BAND = PatternFill("solid", fgColor="F3F4F6")
_FILL_WARN = PatternFill("solid", fgColor="FFFBEB")
_FILL_OK = PatternFill("solid", fgColor="ECFDF5")

_WRAP = Alignment(vertical="top", wrap_text=True)
_TOP = Alignment(vertical="top")

# Verdicts, spelled for a person rather than for the database.
_VERDICT_WORDS = {
    "merge": "Merge into one golden record",
    "keep_survivor": "Keep the nominated row only",
    "keep_all": "Keep all — separate entities",
    "keep_subset": "Keep the ticked rows only",
    "exclude": "Exclude the whole group",
}


def _sheet(wb: Workbook, name: str, title: str, blurb: str):
    ws = wb.create_sheet(name[:31])
    ws.sheet_view.showGridLines = False
    ws["A1"] = title
    ws["A1"].font = _F_TITLE
    ws["A2"] = blurb
    ws["A2"].font = _F_SUB
    ws.row_dimensions[1].height = 22
    return ws


def _headers(ws, row: int, headers: list[tuple[str, int]]) -> None:
    for i, (label, width) in enumerate(headers, start=1):
        c = ws.cell(row=row, column=i, value=label)
        c.font, c.fill, c.alignment = _F_HEAD, _FILL_HEAD, _WRAP
        ws.column_dimensions[get_column_letter(i)].width = width
    ws.row_dimensions[row].height = 28
    ws.freeze_panes = ws.cell(row=row + 1, column=1)


def _row(ws, row: int, values: Iterable[Any], *, band=False, fill=None,
         bold=False) -> None:
    for i, v in enumerate(values, start=1):
        c = ws.cell(row=row, column=i, value=v)
        c.font = _F_GROUP if bold else _F_BODY
        c.alignment = _TOP
        if fill is not None:
            c.fill = fill
        elif band:
            c.fill = _FILL_BAND


def _autofilter(ws, first_row: int, last_row: int, cols: int) -> None:
    if last_row <= first_row:
        return
    ws.auto_filter.ref = (f"A{first_row}:"
                          f"{get_column_letter(cols)}{last_row}")


def conflict_text(cluster: dict) -> str:
    """"Tax Registration Number: A, B, C" for every strong id the group disagrees on.

    The blank is meaningful and is left blank rather than filled with "none": an
    empty cell filters and sorts as the absence it is, whereas a word has to be
    read.
    """
    parts = []
    for c in (cluster.get("id_conflicts") or []):
        vals = [str(v) for v in (c.get("values") or []) if str(v).strip()]
        if len(vals) > 1:
            parts.append(f"{c.get('column')}: {', '.join(vals)}")
    return " · ".join(parts)


def summarise(result: dict) -> dict:
    """Counts for the summary sheet, derived from the SAME clusters the detail
    sheet lists — so the two can never disagree about how many there were."""
    clusters = result.get("clusters") or []
    conflicted = [c for c in clusters if conflict_text(c)]
    decided = [c for c in clusters if (c.get("decision") or {}).get("verdict")]
    return {
        "object": result.get("object") or "",
        "rows_scanned": result.get("rows_scanned"),
        "groups_found": result.get("cluster_count", len(clusters)),
        "groups_listed": len(clusters),
        "records_in_groups": sum(int(c.get("size") or len(c.get("members") or []))
                                 for c in clusters),
        "groups_with_conflicting_ids": len(conflicted),
        "records_in_conflicted_groups": sum(
            int(c.get("size") or len(c.get("members") or [])) for c in conflicted),
        "groups_decided": len(decided),
        "identity_fields": result.get("identity_fields") or [],
        "coverage_note": result.get("coverage_note") or "",
        "sources": result.get("sources") or [],
        "ai_used": bool(result.get("ai_used")),
    }


def _summary(wb: Workbook, s: dict, title: str, generated_at: Any) -> None:
    ws = _sheet(wb, "Summary", "Duplicate suspects",
                "Records the tool believes are the same entity despite different "
                "keys or names — what exact-key de-duplication cannot catch. "
                "Nothing here has been applied to any file; these are candidates "
                "for a decision.")
    rows = [
        ("Conversion", title),
        ("Generated", generated_at),
        ("Object", s["object"]),
        ("Source files", ", ".join(str(x) for x in s["sources"]) or "—"),
        ("Records scanned", s["rows_scanned"]),
        ("", ""),
        ("Suspected groups", s["groups_found"]),
        ("Groups in this file", s["groups_listed"]),
        ("Records inside those groups", s["records_in_groups"]),
        ("Groups already decided", s["groups_decided"]),
        ("", ""),
        ("Groups whose strong IDs DISAGREE", s["groups_with_conflicting_ids"]),
        ("Records in those groups", s["records_in_conflicted_groups"]),
        ("", ""),
        ("Matched on", ", ".join(str(f) for f in s["identity_fields"]) or "—"),
        ("Scan coverage", s["coverage_note"] or "Every scanned record was compared."),
        ("AI adjudication", "Yes" if s["ai_used"] else "No — deterministic scoring only"),
    ]
    ws.column_dimensions["A"].width = 36
    ws.column_dimensions["B"].width = 96
    r = 4
    for label, value in rows:
        if label:
            ws.cell(row=r, column=1, value=label).font = _F_GROUP
            c = ws.cell(row=r, column=2, value=value)
            c.font, c.alignment = _F_BODY, _WRAP
        r += 1

    r += 1
    ws.cell(row=r, column=1, value="How to read this").font = _F_GROUP
    r += 1
    for line in (
        'One row per record, grouped. "Match" is the lowest pairwise confidence '
        'inside the group, so it is the weakest link rather than the best one.',
        '"Conflicting IDs" lists the strong identifiers the group DISAGREES on. '
        'Filter on it: those groups are the ones most likely not to be one entity.',
        'A conflict is advisory and nothing acts on it. A differing tax id usually '
        'means separate legal entities — but it can equally be one source recording '
        'the number wrong, and splitting the group automatically would hide a real '
        'duplicate.',
        'Merging keeps ONE value per column: the survivor first, then the first '
        'non-blank value from the remaining rows. Every other value in that column '
        'is gone from the output.',
        '"Decision" is what is recorded today. Blank means nobody has ruled on the '
        'group and every one of its records still ships.',
    ):
        c = ws.cell(row=r, column=2, value="• " + line)
        c.font, c.alignment = _F_BODY, _WRAP
        ws.row_dimensions[r].height = 30
        r += 1


def _groups(wb: Workbook, clusters: list[dict], identity_fields: list) -> None:
    show = [str(f if isinstance(f, str) else f.get("column")) for f in
            (identity_fields or [])]
    show = [f for f in show if f]
    base = [("Group", 8), ("Match", 9), ("Records", 9), ("Matched on", 26),
            ("Conflicting IDs", 52), ("Decision", 30), ("Row", 8)]
    ws = _sheet(wb, "Groups", "Every suspected group, record by record",
                "Sorted by match confidence, highest first — the same order the "
                "screen uses. Filter 'Conflicting IDs' to the non-blank rows to see "
                "the groups whose strong identifiers disagree.")
    _headers(ws, 4, base + [(f, 30) for f in show])
    r = 5
    for gi, cl in enumerate(clusters, start=1):
        conflict = conflict_text(cl)
        verdict = (cl.get("decision") or {}).get("verdict")
        decision = _VERDICT_WORDS.get(verdict, verdict or "")
        members = cl.get("members") or []
        fill = _FILL_WARN if conflict else (_FILL_OK if verdict else None)
        for mi, m in enumerate(members):
            vals = m.get("values") or {}
            first = mi == 0
            _row(ws, r, [
                gi,
                f"{round(float(cl.get('confidence') or 0) * 100)}%" if first else "",
                int(cl.get("size") or len(members)) if first else "",
                ", ".join(str(x) for x in (cl.get("fields") or [])) if first else "",
                conflict if first else "",
                decision if first else "",
                m.get("row"),
            ] + [vals.get(f, "") for f in show], fill=fill, bold=first)
            r += 1
        # A blank line between groups. Costs one row and is the difference between
        # a list you can read and a wall of 831 rows.
        r += 1
    if not clusters:
        _row(ws, r, ["—", "", "", "", "", "No suspected duplicate groups.", ""])
        r += 1
    _autofilter(ws, 4, r - 1, len(base) + len(show))


def _conflicts(wb: Workbook, clusters: list[dict]) -> None:
    ws = _sheet(wb, "Conflicting IDs", "Groups whose strong identifiers disagree",
                "The subset most likely NOT to be a single entity — and the subset "
                "where merging costs the most, because a merge keeps one value per "
                "column and discards the rest.")
    headers = [("Group", 8), ("Match", 9), ("Records", 9), ("Column", 30),
               ("Values found", 30), ("The differing values", 78)]
    _headers(ws, 4, headers)
    r = 5
    n = 0
    for gi, cl in enumerate(clusters, start=1):
        for c in (cl.get("id_conflicts") or []):
            vals = [str(v) for v in (c.get("values") or []) if str(v).strip()]
            if len(vals) < 2:
                continue
            _row(ws, r, [gi,
                         f"{round(float(cl.get('confidence') or 0) * 100)}%",
                         int(cl.get("size") or len(cl.get("members") or [])),
                         c.get("column"), len(vals), ", ".join(vals)],
                 band=(n % 2 == 1))
            r += 1
            n += 1
    if not n:
        _row(ws, r, ["—", "", "", "", "",
                     "No group disagrees on a strong identifier."])
        r += 1
    _autofilter(ws, 4, r - 1, len(headers))


def build_workbook(*, title: str, generated_at: Any = None,
                   result: dict) -> bytes:
    """The whole scan as a workbook. Dicts in, bytes out — no database, no IO."""
    clusters = list(result.get("clusters") or [])
    wb = Workbook()
    wb.remove(wb.active)
    _summary(wb, summarise(result), title, generated_at or "")
    _groups(wb, clusters, result.get("identity_fields") or [])
    _conflicts(wb, clusters)
    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()
