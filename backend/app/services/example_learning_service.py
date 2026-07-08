"""Learn a conversion's mappings + defaults from a populated example output.

Given a filled-in target template (e.g. the client's gold FBDI file) and the
conversion's source dataset, this infers, per target field:
  * a CONSTANT default  — when the example holds one value for that column, or
  * a source->target MAPPING — the source column whose value SET best overlaps
    the example column's values (set overlap, so it's robust to row reordering
    between the example and the source extract).

The inferred rules are written as MappingSuggestion rows on the conversion (so
Generate Output uses them) and also saved as reusable LearnedMapping reference
standards for the object, so the next conversion of the same object reuses them.
"""
from __future__ import annotations

import logging
import re
from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook

from app.models.conversion import Conversion
from app.models.dataset import Dataset
from app.models.fbdi import FBDIField, FBDITemplate
from app.models.learned import LearnedMapping
from app.models.mapping import MappingSuggestion
from app.parsers import parse_tabular
from app.services.dataset_file_store import materialize_dataset_file

logger = logging.getLogger(__name__)

_MAX_ROWS = 400          # rows of the example/source to compare
_MAP_THRESHOLD = 0.55    # min value-set overlap to accept a source->target map
_BLANKS = {"", "nan", "none", "null", "nat"}


def _norm_key(s: str) -> str:
    return re.sub(r"[^a-z0-9]", "", (s or "").strip().lower().strip("*"))


def _norm_val(v) -> str:
    if v is None:
        return ""
    s = str(v).strip()
    if s.lower() in _BLANKS:
        return ""
    # normalize numbers like "1416.0" -> "1416" so int/float source cols match
    if re.fullmatch(r"-?\d+\.0+", s):
        s = s.split(".")[0]
    return s.lower()


def _read_example(path: Path) -> dict[str, list[str]]:
    """Return {normalized_header_key: [row values]} across all non-instruction
    sheets of a populated template. Header = the row with the most text cells in
    the first 6 rows (matches the FBDI parser's detection)."""
    wb = load_workbook(filename=path, data_only=True, read_only=True)
    out: dict[str, list[str]] = {}
    for sname in wb.sheetnames:
        if "instruction" in sname.lower():
            continue
        ws = wb[sname]
        rows = []
        for i, r in enumerate(ws.iter_rows(values_only=True)):
            rows.append(r)
            if i > _MAX_ROWS:
                break
        if not rows:
            continue
        def score(r):
            return sum(1 for c in r if isinstance(c, str) and c.strip())
        hi = max(range(min(6, len(rows))), key=lambda i: score(rows[i]))
        header = [str(c).strip() if c is not None else "" for c in rows[hi]]
        data = rows[hi + 1:]
        for ci, col in enumerate(header):
            if not col:
                continue
            key = _norm_key(col)
            if not key or key in out:
                continue
            out[key] = [(_norm_val(r[ci]) if ci < len(r) else "") for r in data]
    wb.close()
    return out


async def _upsert_mapping(conversion, field, *, source_column, default_value, confidence, reason,
                          status="approved"):
    existing = await MappingSuggestion.find_one(
        MappingSuggestion.conversion_id == conversion.id,
        MappingSuggestion.target_field_id == field.id,
    )
    payload = {
        "source_column": source_column,
        "default_value": default_value,
        "confidence": float(confidence),
        "reason": reason,
        "status": status,
        "review_required": 0,
        "updated_at": datetime.utcnow(),
    }
    if existing:
        await existing.set(payload)
    else:
        await MappingSuggestion(
            conversion_id=conversion.id, target_field_id=field.id, **payload
        ).insert()


async def _save_reference_standard(target_object, field, *, source_column, default_value, suppress=False):
    """Persist a reusable, OBJECT-LEVEL learned rule so a brand-new conversion of
    the same object auto-applies it on Generate Set — the learning engine's
    ``apply_learned_to_conversion`` re-reads these when auto-mapping.

    Source->target maps are stored as kind='column_mapping' (the exact shape the
    engine reuses: original_value = source column name, resolved_value/target_field
    = target field, keyed by target_object). Constant defaults are recorded
    informationally as 'example_default' (common defaults are also applied in code).
    """
    if not target_object or not field.field_name:
        return
    if suppress:
        # Gold leaves this field blank → reusable rule to keep it blank on every
        # future conversion of this object (overrides aggressive AI mapping).
        kind, category = "suppress_field", "Suppressed (blank in gold)"
        original, resolved, rtype = "(blank)", "", "suppress"
    elif source_column:
        kind, category = "column_mapping", "Column Mapping Alias"
        original, resolved, rtype = source_column, field.field_name, None
    else:
        kind, category = "example_default", "Default Value"
        original, resolved, rtype = "(default)", (default_value or ""), "default"
    existing = await LearnedMapping.find_one(
        LearnedMapping.kind == kind,
        LearnedMapping.target_object == target_object,
        LearnedMapping.target_field == field.field_name,
    )
    doc = {
        "kind": kind,
        "category": category,
        "original_value": str(original),
        "resolved_value": str(resolved),
        "target_object": target_object,
        "target_field": field.field_name,
        "rule_type": rtype,
        "rule_config": {"source_column": source_column, "default_value": default_value},
        "captured_from": "gold example",
        "captured_at": datetime.utcnow(),
    }
    if existing:
        await existing.set(doc)
    else:
        await LearnedMapping(**doc).insert()


async def learn_conversion_from_example(conversion: Conversion, example_path: str | Path) -> dict:
    example_path = Path(example_path)
    template = await FBDITemplate.get(conversion.template_id) if conversion.template_id else None
    if not template:
        return {"error": "conversion has no template"}
    fields = await FBDIField.find(FBDIField.template_id == template.id).sort("+sequence").to_list()

    ex = _read_example(example_path)
    if not ex:
        return {"error": "no readable columns in example file"}

    # Source columns (value lists), read once.
    src_cols: dict[str, list[str]] = {}
    if conversion.dataset_id:
        dataset = await Dataset.get(conversion.dataset_id)
        src_path = await materialize_dataset_file(dataset) if dataset else None
        if src_path is not None:
            src_df = parse_tabular(str(src_path), file_type=dataset.file_type, nrows=_MAX_ROWS)
            for c in src_df.columns:
                src_cols[c] = [_norm_val(v) for v in src_df[c].tolist()]
    # Pre-compute source value sets for overlap scoring.
    src_sets = {c: set(v for v in vals if v) for c, vals in src_cols.items()}

    target_object = template.business_object or conversion.target_object
    mapped, defaulted, skipped, suppressed = [], [], 0, []

    for f in fields:
        ex_vals = ex.get(_norm_key(f.field_name))
        if ex_vals is None:
            skipped += 1
            continue
        nonblank = [v for v in ex_vals if v]
        if not nonblank:
            # Field is a column in the gold example but LEFT BLANK everywhere →
            # the gold says don't populate it. Suppress it (status not_applicable)
            # so it overrides the aggressive AI mapping and stays empty at output,
            # and learn it as a reusable rule for future conversions of this object.
            await _upsert_mapping(conversion, f, source_column=None, default_value=None,
                                  confidence=0.9, reason="blank in gold example",
                                  status="not_applicable")
            await _save_reference_standard(target_object, f, source_column=None,
                                           default_value=None, suppress=True)
            suppressed.append(f.field_name)
            continue
        uniq = set(nonblank)
        # Constant column -> default value (use the original-cased example value).
        if len(uniq) == 1:
            # recover original casing from the raw example (first non-blank)
            val = nonblank[0]
            await _upsert_mapping(conversion, f, source_column=None, default_value=val,
                                  confidence=0.97, reason="constant in example")
            await _save_reference_standard(target_object, f, source_column=None, default_value=val)
            defaulted.append({"field": f.field_name, "value": val})
            continue
        # Otherwise find the source column with the best value-set overlap.
        best_col, best_ratio = None, 0.0
        for c, cset in src_sets.items():
            if not cset:
                continue
            inter = len(uniq & cset)
            ratio = inter / len(uniq)
            if ratio > best_ratio:
                best_ratio, best_col = ratio, c
        if best_col and best_ratio >= _MAP_THRESHOLD:
            await _upsert_mapping(conversion, f, source_column=best_col, default_value=None,
                                  confidence=round(best_ratio, 3), reason=f"value-set match {best_ratio:.0%}")
            await _save_reference_standard(target_object, f, source_column=best_col, default_value=None)
            mapped.append({"field": f.field_name, "source": best_col, "match": round(best_ratio, 2)})
        else:
            skipped += 1

    await conversion.set({"status": "mapped", "updated_at": datetime.utcnow()})
    return {
        "target_object": target_object,
        "mapped_count": len(mapped),
        "default_count": len(defaulted),
        "suppressed_count": len(suppressed),
        "skipped": skipped,
        "mapped": mapped[:60],
        "defaults": defaulted[:60],
        "suppressed": suppressed[:60],
    }
